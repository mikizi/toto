#!/usr/bin/env python3
"""Unit tests for export and validation (no LibreOffice required)."""

from __future__ import annotations

import json
import copy
import tempfile
import unittest
from pathlib import Path

from scripts.export_summary import (
    _cell_int,
    _cell_number,
    _cell_text,
    _knockout_eliminated,
    _movement,
    _reconstructed_summary_leaderboard,
    _read_knockout,
    _read_user_points,
    _read_visible_summary_leaderboard,
    build_export,
    write_export,
)
from scripts.patch_match import clear_match_score, find_match_row, patch_match
import inspect

from scripts.publish_match import (
    _restore_open_match_ids_from_previous,
    close_live_match,
    publish_match,
)
from scripts.validate_export import validate

from scripts.paths import BACKUP_PATH, XLSX_PATH


class TestValidation(unittest.TestCase):
    """Export payload validation."""

    def test_validate_rejects_duplicate_match_ids(self) -> None:
        payload = {
            "version": "test",
            "gamesPlayed": 0,
            "leaderboard": [{"name": "Miki", "champion": "Spain"}],
            "matches": [{"id": 1}, {"id": 1}],
        }
        self.assertIn("matches contain duplicate ids", validate(payload))


class TestCellParsing(unittest.TestCase):
    """Spreadsheet cached-value parsing."""

    def test_cell_number_ignores_excel_errors(self) -> None:
        self.assertIsNone(_cell_number("#N/A"))
        self.assertIsNone(_cell_number("#REF!"))
        self.assertEqual(_cell_number("5"), 5.0)
        self.assertEqual(_cell_number(3), 3.0)

    def test_cell_int_ignores_excel_errors(self) -> None:
        self.assertIsNone(_cell_int("#N/A"))
        self.assertEqual(_cell_int(2.0), 2)

    def test_cell_text_ignores_excel_errors(self) -> None:
        self.assertIsNone(_cell_text("#N/A"))
        self.assertIsNone(_cell_text("#REF!"))
        self.assertEqual(_cell_text("France"), "France")


class TestKnockoutExport(unittest.TestCase):
    """Knockout qualifier parsing from Summary."""

    def test_read_knockout_reads_actual_qualified_teams(self) -> None:
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Summary"
        ws["P4"].value = "Mexico"
        ws["P5"].value = "#N/A"
        ws["R4"].value = "Brazil"
        ws["R42"].value = "France"

        knockout = _read_knockout(ws)

        self.assertEqual(knockout["actual"]["r32"], ["Mexico"])
        self.assertEqual(knockout["actual"]["r16"], ["Brazil"])
        self.assertEqual(knockout["actual"]["champion"], ["France"])
        self.assertEqual(knockout["points"]["r32"], 5)

    def test_knockout_eliminated_uses_api_r32_qualifiers(self) -> None:
        knockout = {
            "actual": {},
            "matches": [
                {
                    "id": 73,
                    "roundId": "r32_match",
                    "apiHome": "France",
                    "apiAway": "Brazil",
                },
                {
                    "id": 74,
                    "roundId": "r32_match",
                    "apiHome": "Germany",
                    "apiAway": "Canada",
                },
            ],
        }
        matches = [
            {"home": "France", "away": "Brazil"},
            {"home": "Germany", "away": "Canada"},
            {"home": "Mexico", "away": "Norway"},
        ]

        eliminated = _knockout_eliminated(knockout, matches)

        self.assertNotIn("r32", eliminated)

        for index in range(14):
            knockout["matches"].append(
                {
                    "id": 75 + index,
                    "roundId": "r32_match",
                    "apiHome": f"Qualified {index}A",
                    "apiAway": f"Qualified {index}B",
                }
            )

        eliminated = _knockout_eliminated(knockout, matches)

        self.assertEqual(eliminated["r32"], ["Mexico", "Norway"])

    def test_knockout_eliminated_uses_api_post_match_loser(self) -> None:
        knockout = {
            "actual": {},
            "matches": [
                {
                    "id": 73,
                    "roundId": "r32_match",
                    "home": "France",
                    "away": "Brazil",
                    "apiState": "post",
                    "apiHomeScore": 2,
                    "apiAwayScore": 1,
                }
            ],
        }

        eliminated = _knockout_eliminated(knockout, [])

        self.assertEqual(eliminated["r16"], ["Brazil"])

    def test_user_points_prefers_cached_summary_total(self) -> None:
        import openpyxl

        wb_data = openpyxl.Workbook()
        ws_data = wb_data.active
        ws_data.title = "Summary"
        ws_data["F79"].value = 123
        wb_data.create_sheet("Calc")
        wb_data["Calc"]["A1"].value = 9

        wb_formulas = openpyxl.Workbook()
        ws_formulas = wb_formulas.active
        ws_formulas.title = "Summary"
        ws_formulas["F79"].value = "=Calc!A1"
        wb_formulas.create_sheet("Calc")

        self.assertEqual(_read_user_points(wb_data, wb_formulas, ws_data, 79), 123)


class TestSummaryLeaderboardParsing(unittest.TestCase):
    """Summary sheet range discovery."""

    def test_visible_leaderboard_reads_past_legacy_row_80(self) -> None:
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Summary"
        raw_by_name = {}
        for row in range(4, 92):
            index = row - 3
            name = f"Player{index}"
            ws[f"A{row}"].value = index
            ws[f"C{row}"].value = index
            ws[f"D{row}"].value = name
            ws[f"E{row}"].value = "France"
            ws[f"F{row}"].value = 100 - index
            raw_by_name[name] = {"id": str(index), "champion": "France", "picks": []}
        ws["C96"].value = "#"
        ws["D96"].value = "Name"

        rows = _read_visible_summary_leaderboard(ws, raw_by_name)

        self.assertEqual(len(rows), 88)
        self.assertEqual(rows[-1]["name"], "Player88")

    def test_visible_leaderboard_uses_excel_real_rank_for_ties(self) -> None:
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Summary"
        ws["A4"].value = 1
        ws["C4"].value = 1
        ws["D4"].value = "LeaderA"
        ws["F4"].value = 10
        ws["A5"].value = 2
        ws["C5"].value = "-"
        ws["D5"].value = "LeaderB"
        ws["F5"].value = 10
        ws["C6"].value = "#"
        ws["D6"].value = "Name"

        rows = _read_visible_summary_leaderboard(
            ws,
            {
                "LeaderA": {"rank": 1, "champion": "France", "picks": []},
                "LeaderB": {"rank": 1, "champion": "Spain", "picks": []},
            },
        )

        self.assertEqual([row["rank"] for row in rows], [1, 1])
        self.assertEqual([row["rankLabel"] for row in rows], ["1", "1"])

    def test_visible_leaderboard_inherits_rank_when_tie_label_has_no_raw_rank(self) -> None:
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Summary"
        ws["A4"].value = 1
        ws["C4"].value = 1
        ws["D4"].value = "LeaderA"
        ws["F4"].value = 10
        ws["A5"].value = 2
        ws["C5"].value = "-"
        ws["D5"].value = "LeaderB"
        ws["F5"].value = 10
        ws["C6"].value = "#"
        ws["D6"].value = "Name"

        rows = _read_visible_summary_leaderboard(ws, {})

        self.assertEqual([row["rank"] for row in rows], [1, 1])
        self.assertEqual([row["rankLabel"] for row in rows], ["1", "1"])

    def test_reconstructed_leaderboard_preserves_excel_real_rank(self) -> None:
        rows = _reconstructed_summary_leaderboard(
            [
                {"name": "LeaderA", "points": 10, "rank": 1, "summaryOrder": 1},
                {"name": "LeaderB", "points": 10, "rank": 1, "summaryOrder": 2},
                {"name": "Third", "points": 8, "rank": 3, "summaryOrder": 3},
            ]
        )

        self.assertEqual([row["rank"] for row in rows], [1, 1, 3])
        self.assertEqual([row["rankLabel"] for row in rows], ["1", "1", "3"])

    def test_reconstructed_leaderboard_uses_rank_eq_fallback_for_ties(self) -> None:
        rows = _reconstructed_summary_leaderboard(
            [
                {"name": "LeaderA", "points": 10, "summaryOrder": 1},
                {"name": "LeaderB", "points": 10, "summaryOrder": 2},
                {"name": "Third", "points": 8, "summaryOrder": 3},
            ]
        )

        self.assertEqual([row["rank"] for row in rows], [1, 1, 3])
        self.assertEqual([row["rankLabel"] for row in rows], ["1", "1", "3"])

    def test_movement_compares_inherited_previous_tie_rank(self) -> None:
        previous = {
            "leaderboard": [
                {"name": "LeaderA", "rank": 1, "rankLabel": "1"},
                {"name": "LeaderB", "rank": 2, "rankLabel": "-"},
            ]
        }
        current = [
            {"name": "LeaderA", "rank": 1, "rankLabel": "1"},
            {"name": "LeaderB", "rank": 1, "rankLabel": "1"},
        ]

        rows = _movement(current, previous)

        self.assertEqual([row["movement"] for row in rows], ["same", "same"])


class TestExportFromXlsx(unittest.TestCase):
    """Read existing xlsx and check export shape."""

    @classmethod
    def setUpClass(cls) -> None:
        if not XLSX_PATH.exists():
            raise unittest.SkipTest("xlsx/Master WorldCup26.xlsx not found")
        cls.payload = build_export(XLSX_PATH)

    def test_build_export_has_leaderboard_and_matches(self) -> None:
        payload = self.payload
        self.assertGreaterEqual(len(payload["leaderboard"]), 6)
        self.assertGreater(len(payload["matches"]), 0)
        self.assertIn("version", payload)
        self.assertIn("gamesPlayed", payload)
        self.assertIn("knockout", payload)

    def test_build_export_has_knockout_rounds(self) -> None:
        payload = self.payload
        knockout = payload["knockout"]
        self.assertEqual(
            [round_def["id"] for round_def in knockout["rounds"]],
            ["r32", "r16", "quarter", "semi", "final", "champion"],
        )
        self.assertEqual(knockout["rounds"][0]["expected"], 32)
        self.assertEqual(knockout["points"]["champion"], 40)

    def test_build_export_has_unique_scheduled_matches(self) -> None:
        payload = self.payload
        match_ids = [match["id"] for match in payload["matches"]]
        self.assertEqual(len(match_ids), 72)
        self.assertEqual(match_ids, list(range(1, 73)))

    def test_leaderboard_entries_have_required_fields(self) -> None:
        payload = self.payload
        entry = payload["leaderboard"][0]
        for key in ("id", "name", "points", "rank", "movement"):
            self.assertIn(key, entry)

    def test_leaderboard_entries_have_champion_picks(self) -> None:
        payload = self.payload
        missing = [entry["name"] for entry in payload["leaderboard"] if not entry["champion"]]
        self.assertEqual(missing, [])

    def test_leaderboard_excludes_test_users(self) -> None:
        payload = self.payload
        names = [e["name"] for e in payload["leaderboard"]]
        self.assertFalse(any(n.lower().startswith("test") for n in names))
        self.assertGreater(len(payload["leaderboard"]), 10)

    def test_leaderboard_has_valid_names_and_score_order(self) -> None:
        payload = self.payload
        names = [entry["name"] for entry in payload["leaderboard"]]
        self.assertFalse(any(name.startswith("#") for name in names))
        points = [entry["points"] for entry in payload["leaderboard"]]
        self.assertEqual(points, sorted(points, reverse=True))

    def test_late_joiners_do_not_get_perfect_historical_pick_points(self) -> None:
        payload = self.payload
        late_joiners = [
            entry for entry in payload["leaderboard"] if entry["name"].strip().startswith("N_")
        ]
        if not late_joiners:
            self.skipTest("workbook has no late joiners")

        historical_points = [
            pick["points"]
            for entry in late_joiners
            for pick in entry["picks"]
            if pick["matchId"] <= 24
        ]
        self.assertTrue(historical_points)
        self.assertTrue(all(points is None for points in historical_points))

    def test_validate_latest_export(self) -> None:
        errors = validate(self.payload)
        self.assertEqual(errors, [], msg="; ".join(errors))

    def test_export_keeps_open_live_matches_even_after_start_score(self) -> None:
        base = self.payload
        played_row = base["matches"][0]
        unplayed_row = base["matches"][1]
        previous = {
            **base,
            "broadcast": {
                "mode": "manual",
                "openMatchIds": [played_row["id"], unplayed_row["id"]],
                "suppressAuto": False,
                "autoPilot": True,
            },
        }
        payload = build_export(XLSX_PATH, previous)
        self.assertEqual(
            payload["broadcast"]["openMatchIds"],
            [played_row["id"], unplayed_row["id"]],
        )

    def test_close_live_match_removes_only_finalized_match(self) -> None:
        payload = {"broadcast": {"openMatchIds": [4, 5], "mode": "manual", "suppressAuto": False}}
        close_live_match(payload, 4)
        self.assertEqual(payload["broadcast"]["openMatchIds"], [5])

    def test_publish_match_default_keeps_live_open(self) -> None:
        default = inspect.signature(publish_match).parameters["close_live"].default
        self.assertIs(default, False)

    def test_restore_open_match_ids_from_previous(self) -> None:
        base = self.payload
        previous = {
            **base,
            "broadcast": {
                "mode": "manual",
                "openMatchIds": [2],
                "suppressAuto": True,
                "autoPilot": False,
            },
        }
        payload = copy.deepcopy(self.payload)
        payload["broadcast"] = {"openMatchIds": [2], "mode": "manual", "suppressAuto": True}
        close_live_match(payload, 2)
        self.assertEqual(payload["broadcast"]["openMatchIds"], [])
        _restore_open_match_ids_from_previous(payload, previous)
        self.assertEqual(payload["broadcast"]["openMatchIds"], [2])

    def test_write_export_roundtrip(self) -> None:
        payload = self.payload
        with tempfile.TemporaryDirectory() as tmp:
            latest = Path(tmp) / "latest.json"
            write_export(
                payload,
                latest_path=latest,
                version_path=Path(tmp) / "v.json",
                versions_dir=Path(tmp) / "versions",
            )
            loaded = json.loads(latest.read_text(encoding="utf-8"))
            self.assertEqual(len(loaded["leaderboard"]), len(payload["leaderboard"]))


class TestPatchMatch(unittest.TestCase):
    """Patch match on a temp copy."""

    @classmethod
    def setUpClass(cls) -> None:
        if not BACKUP_PATH.exists():
            raise unittest.SkipTest("simulation backup not found")

    def test_find_match_row_match_1(self) -> None:
        import openpyxl
        import shutil

        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "test.xlsx"
            shutil.copy2(BACKUP_PATH, path)
            wb = openpyxl.load_workbook(path)
            row = find_match_row(wb["Summary"], 1)
            self.assertEqual(wb["Summary"][f"K{row}"].value, "Mexico-South Africa")

    def test_patch_writes_lm_cells(self) -> None:
        import openpyxl
        import shutil

        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "test.xlsx"
            shutil.copy2(BACKUP_PATH, path)
            teams, home, away = patch_match(1, 2, 1, path)
            self.assertIn("Mexico", teams)
            self.assertEqual((home, away), (2, 1))
            wb = openpyxl.load_workbook(path, data_only=True)
            row = find_match_row(wb["Summary"], 1)
            self.assertEqual(wb["Summary"][f"L{row}"].value, 2)
            self.assertEqual(wb["Summary"][f"M{row}"].value, 1)

    def test_clear_match_score_clears_lm_cells(self) -> None:
        import openpyxl
        import shutil

        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "test.xlsx"
            shutil.copy2(BACKUP_PATH, path)
            patch_match(1, 2, 1, path)
            clear_match_score(1, path)
            wb = openpyxl.load_workbook(path, data_only=True)
            row = find_match_row(wb["Summary"], 1)
            self.assertIsNone(wb["Summary"][f"L{row}"].value)
            self.assertIsNone(wb["Summary"][f"M{row}"].value)

    def test_export_scores_from_picks_without_cached_formulas(self) -> None:
        import openpyxl
        import shutil

        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "test.xlsx"
            shutil.copy2(BACKUP_PATH, path)
            patch_match(1, 1, 0, path)

            wb = openpyxl.load_workbook(path)
            ws = wb["Summary"]
            for row in range(4, 81):
                ws[f"D{row}"].value = None
            for row in range(79, 85):
                ws[f"E{row}"].value = "#N/A"
                ws[f"F{row}"].value = "#N/A"
                ws[f"G{row}"].value = "#N/A"
            wb.save(path)

            payload = build_export(path)
            by_name = {entry["name"]: entry for entry in payload["leaderboard"]}
            self.assertEqual(by_name["Miki_Ziso"]["champion"], "Brazil")
            self.assertEqual(by_name["Miki_Ziso"]["points"], 5.0)
            self.assertEqual(by_name["Nir2"]["points"], 0.0)
            self.assertEqual(by_name["Nir3"]["points"], 3.0)


if __name__ == "__main__":
    unittest.main()
