#!/usr/bin/env python3
"""Unit tests for knockout schedule, scoring, and live points."""

from __future__ import annotations

import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

import openpyxl

from scripts.knockout import (
    KNOCKOUT_POINTS,
    apply_live_knockout_points,
    normalize_knockout_state,
    qualifier_target_for_match,
    sync_knockout_fixtures_from_espn,
    update_scoring_constants,
    validate_knockout_fixture_lock,
)


class TestKnockoutState(unittest.TestCase):
    def test_normalize_keeps_autofilled_fixture_unlocked(self) -> None:
        state = normalize_knockout_state(
            {
                "matches": [
                    {"id": 89, "home": "France", "away": "Brazil", "isLocked": False}
                ]
            }
        )
        match = next(item for item in state["matches"] if item["id"] == 89)

        self.assertEqual(match["home"], "France")
        self.assertEqual(match["away"], "Brazil")
        self.assertFalse(match["isLocked"])
        self.assertFalse(state["scoringApplied"]["r32"])

    def test_qualifier_target_maps_winner_to_next_round_cell(self) -> None:
        target = qualifier_target_for_match(73)

        self.assertEqual(target, ("r16", 0))

    def test_espn_sync_locks_real_fixture_teams(self) -> None:
        state = normalize_knockout_state()
        payload = {
            "events": [
                {
                    "id": "760486",
                    "date": "2026-06-28T19:00Z",
                    "competitions": [
                        {
                            "competitors": [
                                {"homeAway": "home", "team": {"displayName": "South Africa"}, "score": "0"},
                                {"homeAway": "away", "team": {"displayName": "Canada"}, "score": "0"},
                            ]
                        }
                    ],
                }
            ]
        }

        with patch("scripts.knockout.fetch_scoreboard", return_value=payload):
            stats = sync_knockout_fixtures_from_espn(state)

        match = next(item for item in state["matches"] if item["id"] == 73)
        self.assertEqual(stats["matchedCount"], 1)
        self.assertEqual(match["home"], "South Africa")
        self.assertEqual(match["away"], "Canada")
        self.assertEqual(match["apiEventId"], "760486")
        self.assertEqual(match["apiHomeScore"], 0)
        self.assertEqual(match["apiAwayScore"], 0)
        self.assertTrue(match["isLocked"])

    def test_espn_sync_normalizes_placeholders(self) -> None:
        state = normalize_knockout_state()
        payload = {
            "events": [
                {
                    "id": "760489",
                    "date": "2026-06-29T20:30Z",
                    "competitions": [
                        {
                            "competitors": [
                                {"homeAway": "home", "team": {"displayName": "Germany"}, "score": "0"},
                                {"homeAway": "away", "team": {"displayName": "Third Place Group A/B/C/D/F"}, "score": "0"},
                            ]
                        }
                    ],
                }
            ]
        }

        with patch("scripts.knockout.fetch_scoreboard", return_value=payload):
            sync_knockout_fixtures_from_espn(state)

        match = next(item for item in state["matches"] if item["id"] == 74)
        self.assertEqual(match["home"], "Germany")
        self.assertEqual(match["away"], "Best 3rd A/B/C/D/F")
        self.assertFalse(match["isLocked"])

    def test_lock_validation_rejects_placeholders_and_api_mismatch(self) -> None:
        match = {"apiHome": "South Africa", "apiAway": "Canada"}

        with self.assertRaisesRegex(ValueError, "placeholder"):
            validate_knockout_fixture_lock(match, "Winner Group A", "Canada")

        with self.assertRaisesRegex(ValueError, "placeholder"):
            validate_knockout_fixture_lock(match, "Semifinal 1 Winner", "Canada")

        validate_knockout_fixture_lock(match, "Brazil", "Canada")
        validate_knockout_fixture_lock(match, "South Africa", "Canada")


class TestKnockoutScoring(unittest.TestCase):
    def test_scoring_constants_match_written_rules(self) -> None:
        self.assertEqual(
            KNOCKOUT_POINTS,
            {
                "r32": 5,
                "r16": 8,
                "quarter": 18,
                "semi": 25,
                "final": 30,
                "champion": 40,
            },
        )

    def test_update_scoring_constants_updates_values_and_formula_refs(self) -> None:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Calc"
        ws["J149"] = '=IF(Summary!R22=$C149,$E$178,0)'
        ws["J159"] = '=IF(Summary!R32=$C159,$E$179,0)'
        ws["J165"] = '=IF(Summary!R38=$C165,$E$180,0)'
        ws["J169"] = '=IF(Summary!R42=$C169,$E$181,0)'

        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "knockout.xlsx"
            wb.save(path)
            update_scoring_constants(path)
            updated = openpyxl.load_workbook(path, data_only=False)["Calc"]

        self.assertEqual(updated["E177"].value, 5)
        self.assertEqual(updated["E178"].value, 8)
        self.assertEqual(updated["E179"].value, 18)
        self.assertEqual(updated["E180"].value, 25)
        self.assertEqual(updated["E181"].value, 30)
        self.assertEqual(updated["E182"].value, 40)
        self.assertIn("$E$179", updated["J149"].value)
        self.assertIn("$E$180", updated["J159"].value)
        self.assertIn("$E$181", updated["J165"].value)
        self.assertIn("$E$182", updated["J169"].value)


class TestLiveKnockoutPoints(unittest.TestCase):
    def test_live_tie_awards_no_temporary_points(self) -> None:
        payload = {
            "knockout": {
                "matches": [
                    {
                        "id": 73,
                        "roundId": "r32",
                        "home": "France",
                        "away": "Brazil",
                        "homeScore": 1,
                        "awayScore": 1,
                        "isLive": True,
                        "isScoring": True,
                    }
                ]
            },
            "leaderboard": [
                {
                    "name": "A",
                    "points": 10,
                    "knockoutPicks": [
                        {
                            "roundId": "r16",
                            "pointsPerTeam": 8,
                            "teams": [{"team": "France", "isCorrect": False}],
                        }
                    ],
                }
            ],
        }

        apply_live_knockout_points(payload)

        self.assertEqual(payload["leaderboard"][0]["points"], 10)
        self.assertNotIn("liveKnockoutPoints", payload["leaderboard"][0])

    def test_live_leader_awards_temporary_points_and_resorts(self) -> None:
        payload = {
            "knockout": {
                "matches": [
                    {
                        "id": 73,
                        "roundId": "r32",
                        "home": "France",
                        "away": "Brazil",
                        "homeScore": 2,
                        "awayScore": 1,
                        "isLive": True,
                        "isScoring": True,
                    }
                ]
            },
            "leaderboard": [
                {
                    "name": "B",
                    "points": 12,
                    "knockoutPicks": [
                        {
                            "roundId": "r16",
                            "pointsPerTeam": 8,
                            "teams": [{"team": "Brazil", "isCorrect": False}],
                        }
                    ],
                },
                {
                    "name": "A",
                    "points": 10,
                    "knockoutPicks": [
                        {
                            "roundId": "r16",
                            "pointsPerTeam": 8,
                            "teams": [{"team": "France", "isCorrect": False}],
                        }
                    ],
                },
            ],
        }

        apply_live_knockout_points(payload)

        self.assertEqual(payload["leaderboard"][0]["name"], "A")
        self.assertEqual(payload["leaderboard"][0]["points"], 18)
        self.assertEqual(payload["leaderboard"][0]["liveKnockoutPoints"], 8)
        self.assertTrue(payload["leaderboard"][0]["knockoutPicks"][0]["teams"][0]["isLiveCorrect"])


if __name__ == "__main__":
    unittest.main()
