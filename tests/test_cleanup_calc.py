#!/usr/bin/env python3
"""Tests for Calc sheet cleanup."""

from __future__ import annotations

import shutil
import tempfile
import unittest
from pathlib import Path

import openpyxl

from scripts.cleanup_calc import cleanup_calc
from scripts.libreoffice_recalc import recalc
from scripts.paths import BACKUP_PATH, XLSX_PATH


class TestCleanupCalc(unittest.TestCase):
    @classmethod
    def setUpClass(cls) -> None:
        if not XLSX_PATH.exists():
            raise unittest.SkipTest("xlsx not found")
        if not shutil.which("soffice"):
            raise unittest.SkipTest("LibreOffice not installed")

    def test_cleanup_zeros_test_users_after_recalc(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            work = Path(tmp) / "cleanup.xlsx"
            shutil.copy2(XLSX_PATH, work)
            wb = openpyxl.load_workbook(work)
            cleanup_calc(wb)
            wb.save(work)
            recalc(work)

            wb = openpyxl.load_workbook(work, data_only=True)
            ws = wb["Summary"]
            for name in ("test25", "test13", "MikiZiso3"):
                for row in range(79, 149):
                    if ws[f"D{row}"].value == name:
                        points = ws[f"F{row}"].value
                        self.assertEqual(
                            float(points or 0),
                            0.0,
                            msg=f"{name} should have 0 pts after cleanup",
                        )
                        break


if __name__ == "__main__":
    unittest.main()
