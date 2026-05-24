#!/usr/bin/env python3
"""Integration tests: reset → recalc → patch → recalc → export (needs LibreOffice)."""

from __future__ import annotations

import shutil
import subprocess
import sys
import tempfile
import unittest
from pathlib import Path

import openpyxl

from scripts.paths import BACKUP_PATH

KNOCKOUT_HEADERS = {"Quarterfinals", "Semi", "Final", "Winner", "Round of 16", "Round of 32"}
REAL_USERS = ["MikiZiso3", "MikiZiso2", "Miki_Ziso", "Nir1", "Nir2", "Nir3"]


def _soffice_available() -> bool:
    return shutil.which("soffice") is not None


def _recalc(path: Path) -> None:
    result = subprocess.run(
        [
            "soffice",
            "--headless",
            "--invisible",
            "--nodefault",
            "--nologo",
            "--norestore",
            "--convert-to",
            "xlsx",
            "--outdir",
            str(path.parent),
            str(path),
        ],
        capture_output=True,
        text=True,
        timeout=120,
    )
    if result.returncode != 0:
        raise RuntimeError(result.stderr or "LibreOffice recalc failed")


def _reset_day_zero(path: Path) -> None:
    wb = openpyxl.load_workbook(path)
    ws = wb["Summary"]
    for row in range(4, 120):
        if ws[f"J{row}"].value and ws[f"K{row}"].value and "-" in str(ws[f"K{row}"].value):
            ws[f"L{row}"].value = None
            ws[f"M{row}"].value = None
        for col in "PR":
            val = ws[f"{col}{row}"].value
            if val is not None and str(val).strip() not in KNOCKOUT_HEADERS:
                ws[f"{col}{row}"].value = None
    wb.save(path)


def _read_real_user_points(path: Path) -> dict[str, float]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Summary"]
    points: dict[str, float] = {}
    for row in range(79, 149):
        name = ws[f"D{row}"].value
        if name in REAL_USERS:
            val = ws[f"F{row}"].value
            points[str(name)] = float(val) if val is not None else 0.0
    return points


def _predictions_intact(path: Path) -> bool:
    wb = openpyxl.load_workbook(path, data_only=True)
    u = wb["001_MikiZiso3"]
    return u["F7"].value == 1 and u["G7"].value == 0


class TestPipelineIntegration(unittest.TestCase):
    """End-to-end pipeline on a temp copy of the backup xlsx."""

    @classmethod
    def setUpClass(cls) -> None:
        if not BACKUP_PATH.exists():
            raise unittest.SkipTest("xlsx/Master WorldCup26.simulation-backup.xlsx not found")
        if not _soffice_available():
            raise unittest.SkipTest("soffice (LibreOffice) not installed")

    def setUp(self) -> None:
        self._tmpdir = tempfile.TemporaryDirectory()
        self.work = Path(self._tmpdir.name) / "pipeline.xlsx"
        shutil.copy2(BACKUP_PATH, self.work)

    def tearDown(self) -> None:
        self._tmpdir.cleanup()

    def test_day_zero_real_users_all_zero(self) -> None:
        _reset_day_zero(self.work)
        self.assertTrue(_predictions_intact(self.work))
        _recalc(self.work)
        pts = _read_real_user_points(self.work)
        self.assertEqual(len(pts), 6)
        for name, score in pts.items():
            self.assertEqual(score, 0.0, msg=f"{name} should be 0 at day zero")

    def test_patch_match_1_updates_scoreboard(self) -> None:
        from scripts.patch_match import patch_match

        _reset_day_zero(self.work)
        _recalc(self.work)
        before = _read_real_user_points(self.work)

        patch_match(1, 1, 0, self.work)
        _recalc(self.work)
        after = _read_real_user_points(self.work)

        self.assertEqual(before["MikiZiso3"], 0.0)
        self.assertGreater(after["MikiZiso3"], before["MikiZiso3"])
        self.assertAlmostEqual(after["MikiZiso3"], 5.0, places=2)
        self.assertEqual(after["Nir1"], 0.0)
        self.assertGreater(after["Nir3"], 0.0)

    def test_export_after_patch_passes_validation(self) -> None:
        from scripts.export_summary import build_export
        from scripts.patch_match import patch_match
        from scripts.validate_export import validate

        _reset_day_zero(self.work)
        patch_match(1, 1, 0, self.work)
        _recalc(self.work)
        payload = build_export(self.work)
        errors = validate(payload)
        self.assertEqual(errors, [])
        self.assertEqual(payload["gamesPlayed"], 1)
        self.assertIsNotNone(payload["lastResult"])
        self.assertEqual(payload["lastResult"]["matchId"], 1)


if __name__ == "__main__":
    unittest.main()
