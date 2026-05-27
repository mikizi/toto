#!/usr/bin/env python3
"""Reset tournament actual scores to zero played games (pre-tournament state)."""

from __future__ import annotations

import argparse
import shutil
from pathlib import Path
from typing import Any

import openpyxl

from scripts.paths import BACKUP_PATH, XLSX_PATH
from scripts.cleanup_calc import cleanup_calc
from scripts.libreoffice_recalc import recalc

SUMMARY = "Summary"
KNOCKOUT_HEADERS = {"Quarterfinals", "Semi", "Final", "Winner", "Round of 16", "Round of 32"}
MATCH_ROW_START = 4
MATCH_ROW_END = 120


def _is_match_row_summary(ws: openpyxl.worksheet.worksheet.Worksheet, row: int) -> bool:
    teams = ws[f"K{row}"].value
    match_id = ws[f"J{row}"].value
    return match_id is not None and teams is not None and "-" in str(teams)


def _clear_summary(ws: openpyxl.worksheet.worksheet.Worksheet) -> int:
    cleared = 0
    for row in range(MATCH_ROW_START, MATCH_ROW_END):
        if _is_match_row_summary(ws, row):
            ws[f"L{row}"].value = None
            ws[f"M{row}"].value = None
            cleared += 1

        p_val = ws[f"P{row}"].value
        if p_val is not None and str(p_val).strip() not in KNOCKOUT_HEADERS:
            ws[f"P{row}"].value = None

        r_val = ws[f"R{row}"].value
        if r_val is not None and str(r_val).strip() not in KNOCKOUT_HEADERS:
            ws[f"R{row}"].value = None

    return cleared


def reset_scores(
    xlsx_path: Path = XLSX_PATH,
    backup: bool = True,
    recalc_after: bool = True,
) -> None:
    """Clear actual results only (Summary L/M and knockout P/R). Keeps predictions."""
    if backup:
        shutil.copy2(xlsx_path, BACKUP_PATH)
        print(f"Backup saved → {BACKUP_PATH}")

    wb = openpyxl.load_workbook(xlsx_path)
    summary_cleared = _clear_summary(wb[SUMMARY])
    sheet_users, test_rows = cleanup_calc(wb)
    wb.save(xlsx_path)
    print(f"Cleared {summary_cleared} Summary match scores (L/M)")
    print(f"Calc cleanup: {sheet_users} sheet users, {test_rows} test rows zeroed")
    print("Predictions on user sheets were NOT touched (F/G are picks, not results).")
    if recalc_after:
        recalc(xlsx_path)
        print(f"Recalculated → {xlsx_path}")
    else:
        print("Run: python scripts/libreoffice_recalc.py && python scripts/export_summary.py")


def reset_tournament(
    xlsx_path: Path = XLSX_PATH,
    *,
    backup: bool = True,
) -> dict[str, Any]:
    """Clear all results in xlsx, recalc, export site JSON, and clear live broadcast."""
    from scripts.export_summary import build_export, write_export
    from scripts.update_broadcast import apply_broadcast_update
    from scripts.validate_export import validate

    reset_scores(xlsx_path=xlsx_path, backup=backup, recalc_after=True)
    payload = build_export(xlsx_path, previous=None)
    payload = apply_broadcast_update(
        payload,
        open_match_ids=[],
        suppress_auto=False,
        mode="auto",
        clear_manual=True,
    )
    write_export(payload)
    errors = validate(payload)
    if errors:
        raise RuntimeError(f"Export validation failed: {errors}")
    print(
        f"Tournament reset: {payload['gamesPlayed']} games played, "
        f"version {payload['version']}"
    )
    return payload


def main() -> int:
    parser = argparse.ArgumentParser(description="Reset tournament scores and site data")
    parser.add_argument(
        "--export-only",
        action="store_true",
        help="Only rebuild latest.json from current xlsx (no score clear)",
    )
    parser.add_argument(
        "--no-backup",
        action="store_true",
        help="Skip copying xlsx to simulation-backup before clearing",
    )
    args = parser.parse_args()
    backup = not args.no_backup
    if args.export_only:
        from scripts.export_summary import build_export, write_export

        payload = build_export(XLSX_PATH)
        write_export(payload)
        print(f"Exported version {payload['version']}")
        return 0
    reset_tournament(backup=backup)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
