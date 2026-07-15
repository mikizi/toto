"""Knockout schedule, state, and workbook helpers."""

from __future__ import annotations

from copy import deepcopy
from datetime import datetime, timezone
from pathlib import Path
import re
from typing import Any

import openpyxl

from scripts.espn_scores import EspnMatch, fetch_scoreboard, normalize_team_name, parse_espn_events
from scripts.live_state import parse_iso
from scripts.paths import XLSX_PATH

SUMMARY = "Summary"

KNOCKOUT_POINTS = {
    "r32": 5,
    "r16": 8,
    "quarter": 18,
    "semi": 25,
    "final": 30,
    "champion": 40,
}

KNOCKOUT_ROUNDS = (
    {"id": "r32", "label": "Round of 32", "expected": 32, "points": 5, "range": "P4:P35"},
    {"id": "r16", "label": "Round of 16", "expected": 16, "points": 8, "range": "R4:R19"},
    {"id": "quarter", "label": "Quarter-finals", "expected": 8, "points": 18, "range": "R22:R29"},
    {"id": "semi", "label": "Semi-finals", "expected": 4, "points": 25, "range": "R32:R35"},
    {"id": "final", "label": "Final", "expected": 2, "points": 30, "range": "R38:R39"},
    {"id": "champion", "label": "Winner", "expected": 1, "points": 40, "range": "R42:R42"},
)

ROUND_ACTUAL_CELLS = {
    "r32": [f"P{row}" for row in range(4, 36)],
    "r16": [f"R{row}" for row in range(4, 20)],
    "quarter": [f"R{row}" for row in range(22, 30)],
    "semi": [f"R{row}" for row in range(32, 36)],
    "final": [f"R{row}" for row in range(38, 40)],
    "champion": ["R42"],
}

PLAYER_KNOCKOUT_PICK_CELLS = {
    "r32": [
        "BL18", "BL10", "BL22", "BL42", "BL14", "BL46", "BL50", "BL54",
        "BL34", "BL38", "BL26", "BL30", "BL66", "BL58", "BL70", "BL62",
        "BL19", "BL11", "BL23", "BL43", "BL15", "BL47", "BL51", "BL55",
        "BL35", "BL39", "BL27", "BL31", "BL67", "BL59", "BL71", "BL63",
    ],
    "r16": [
        "BS12", "BS20", "BS44", "BS52", "BS28", "BS36", "BS60", "BS68",
        "BS13", "BS21", "BS45", "BS53", "BS29", "BS37", "BS61", "BS69",
    ],
    "quarter": ["BZ16", "BZ32", "BZ48", "BZ64", "BZ17", "BZ33", "BZ49", "BZ65"],
    "semi": ["CG23", "CG55", "CG24", "CG56"],
    "final": ["CN37", "CN38"],
    "champion": ["CM47"],
}

ESPN_KNOCKOUT_DATES = "20260628-20260719"

ESPN_TEAM_LABEL_ALIASES = {
    "Bosnia-Herzegovina": "Bosnia and Herzegovina",
    "Congo DR": "DR Congo",
    "Czechia": "Czech Republic",
    "South Korea": "Korea Republic",
    "Türkiye": "Turkey",
}

# Dates are the workbook/public schedule dates in UTC-ish site display time.
KNOCKOUT_SCHEDULE = (
    {"id": 73, "roundId": "r32_match", "roundLabel": "Round of 32", "kickoffAt": "2026-06-28T19:00:00+00:00", "homeSlot": "Runner-up Group A", "awaySlot": "Runner-up Group B", "nextMatchId": 90},
    {"id": 74, "roundId": "r32_match", "roundLabel": "Round of 32", "kickoffAt": "2026-06-29T20:30:00+00:00", "homeSlot": "Winner Group E", "awaySlot": "Best 3rd A/B/C/D/F", "nextMatchId": 89},
    {"id": 75, "roundId": "r32_match", "roundLabel": "Round of 32", "kickoffAt": "2026-06-30T01:00:00+00:00", "homeSlot": "Winner Group F", "awaySlot": "Runner-up Group C", "nextMatchId": 90},
    {"id": 76, "roundId": "r32_match", "roundLabel": "Round of 32", "kickoffAt": "2026-06-29T17:00:00+00:00", "homeSlot": "Winner Group C", "awaySlot": "Runner-up Group F", "nextMatchId": 91},
    {"id": 77, "roundId": "r32_match", "roundLabel": "Round of 32", "kickoffAt": "2026-06-30T21:00:00+00:00", "homeSlot": "Winner Group I", "awaySlot": "Best 3rd C/D/F/G/H", "nextMatchId": 89},
    {"id": 78, "roundId": "r32_match", "roundLabel": "Round of 32", "kickoffAt": "2026-06-30T17:00:00+00:00", "homeSlot": "Runner-up Group E", "awaySlot": "Runner-up Group I", "nextMatchId": 91},
    {"id": 79, "roundId": "r32_match", "roundLabel": "Round of 32", "kickoffAt": "2026-07-01T01:00:00+00:00", "homeSlot": "Winner Group A", "awaySlot": "Best 3rd C/E/F/H/I", "nextMatchId": 92},
    {"id": 80, "roundId": "r32_match", "roundLabel": "Round of 32", "kickoffAt": "2026-07-01T16:00:00+00:00", "homeSlot": "Winner Group L", "awaySlot": "Best 3rd E/H/I/J/K", "nextMatchId": 92},
    {"id": 81, "roundId": "r32_match", "roundLabel": "Round of 32", "kickoffAt": "2026-07-02T00:00:00+00:00", "homeSlot": "Winner Group D", "awaySlot": "Best 3rd B/E/F/I/J", "nextMatchId": 94},
    {"id": 82, "roundId": "r32_match", "roundLabel": "Round of 32", "kickoffAt": "2026-07-01T20:00:00+00:00", "homeSlot": "Winner Group G", "awaySlot": "Best 3rd A/E/H/I/J", "nextMatchId": 94},
    {"id": 83, "roundId": "r32_match", "roundLabel": "Round of 32", "kickoffAt": "2026-07-02T23:00:00+00:00", "homeSlot": "Runner-up Group K", "awaySlot": "Runner-up Group L", "nextMatchId": 93},
    {"id": 84, "roundId": "r32_match", "roundLabel": "Round of 32", "kickoffAt": "2026-07-02T19:00:00+00:00", "homeSlot": "Winner Group H", "awaySlot": "Runner-up Group J", "nextMatchId": 93},
    {"id": 85, "roundId": "r32_match", "roundLabel": "Round of 32", "kickoffAt": "2026-07-03T03:00:00+00:00", "homeSlot": "Winner Group B", "awaySlot": "Best 3rd E/F/G/I/J", "nextMatchId": 96},
    {"id": 86, "roundId": "r32_match", "roundLabel": "Round of 32", "kickoffAt": "2026-07-03T22:00:00+00:00", "homeSlot": "Winner Group J", "awaySlot": "Runner-up Group H", "nextMatchId": 95},
    {"id": 87, "roundId": "r32_match", "roundLabel": "Round of 32", "kickoffAt": "2026-07-04T01:30:00+00:00", "homeSlot": "Winner Group K", "awaySlot": "Best 3rd D/E/I/J/L", "nextMatchId": 96},
    {"id": 88, "roundId": "r32_match", "roundLabel": "Round of 32", "kickoffAt": "2026-07-03T18:00:00+00:00", "homeSlot": "Runner-up Group D", "awaySlot": "Runner-up Group G", "nextMatchId": 95},
    {"id": 89, "roundId": "r16_match", "roundLabel": "Round of 16", "kickoffAt": "2026-07-04T21:00:00+00:00", "homeSlot": "Winner 74", "awaySlot": "Winner 77", "nextMatchId": 97},
    {"id": 90, "roundId": "r16_match", "roundLabel": "Round of 16", "kickoffAt": "2026-07-04T17:00:00+00:00", "homeSlot": "Winner 73", "awaySlot": "Winner 75", "nextMatchId": 97},
    {"id": 91, "roundId": "r16_match", "roundLabel": "Round of 16", "kickoffAt": "2026-07-05T20:00:00+00:00", "homeSlot": "Winner 76", "awaySlot": "Winner 78", "nextMatchId": 99},
    {"id": 92, "roundId": "r16_match", "roundLabel": "Round of 16", "kickoffAt": "2026-07-06T00:00:00+00:00", "homeSlot": "Winner 79", "awaySlot": "Winner 80", "nextMatchId": 99},
    {"id": 93, "roundId": "r16_match", "roundLabel": "Round of 16", "kickoffAt": "2026-07-06T19:00:00+00:00", "homeSlot": "Winner 83", "awaySlot": "Winner 84", "nextMatchId": 98},
    {"id": 94, "roundId": "r16_match", "roundLabel": "Round of 16", "kickoffAt": "2026-07-07T00:00:00+00:00", "homeSlot": "Winner 81", "awaySlot": "Winner 82", "nextMatchId": 98},
    {"id": 95, "roundId": "r16_match", "roundLabel": "Round of 16", "kickoffAt": "2026-07-07T16:00:00+00:00", "homeSlot": "Winner 86", "awaySlot": "Winner 88", "nextMatchId": 100},
    {"id": 96, "roundId": "r16_match", "roundLabel": "Round of 16", "kickoffAt": "2026-07-07T20:00:00+00:00", "homeSlot": "Winner 85", "awaySlot": "Winner 87", "nextMatchId": 100},
    {"id": 97, "roundId": "quarter_match", "roundLabel": "Quarter-finals", "kickoffAt": "2026-07-09T20:00:00+00:00", "homeSlot": "Winner 89", "awaySlot": "Winner 90", "nextMatchId": 101},
    {"id": 98, "roundId": "quarter_match", "roundLabel": "Quarter-finals", "kickoffAt": "2026-07-10T19:00:00+00:00", "homeSlot": "Winner 93", "awaySlot": "Winner 94", "nextMatchId": 101},
    {"id": 99, "roundId": "quarter_match", "roundLabel": "Quarter-finals", "kickoffAt": "2026-07-11T21:00:00+00:00", "homeSlot": "Winner 91", "awaySlot": "Winner 92", "nextMatchId": 102},
    {"id": 100, "roundId": "quarter_match", "roundLabel": "Quarter-finals", "kickoffAt": "2026-07-12T01:00:00+00:00", "homeSlot": "Winner 95", "awaySlot": "Winner 96", "nextMatchId": 102},
    {"id": 101, "roundId": "semi_match", "roundLabel": "Semi-finals", "kickoffAt": "2026-07-14T19:00:00+00:00", "homeSlot": "Winner 97", "awaySlot": "Winner 98", "nextMatchId": 104},
    {"id": 102, "roundId": "semi_match", "roundLabel": "Semi-finals", "kickoffAt": "2026-07-15T19:00:00+00:00", "homeSlot": "Winner 99", "awaySlot": "Winner 100", "nextMatchId": 104},
    {"id": 104, "roundId": "final_match", "roundLabel": "Final", "kickoffAt": "2026-07-19T19:00:00+00:00", "homeSlot": "Winner 101", "awaySlot": "Winner 102", "nextMatchId": None},
)

NEXT_MATCH_SIDES = {
    73: "home", 75: "away", 74: "home", 77: "away",
    76: "home", 78: "away", 79: "home", 80: "away",
    83: "home", 84: "away", 81: "home", 82: "away",
    86: "home", 88: "away", 85: "home", 87: "away",
    89: "home", 90: "away", 93: "home", 94: "away",
    91: "home", 92: "away", 95: "home", 96: "away",
    97: "home", 98: "away", 99: "home", 100: "away",
    101: "home", 102: "away",
}

QUALIFIER_ROUND_BY_MATCH_ROUND = {
    "r32_match": "r16",
    "r16_match": "quarter",
    "quarter_match": "semi",
    "semi_match": "final",
    "final_match": "champion",
}


def _cell_text(value: object) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text or text.startswith("#"):
        return None
    return text


def normalize_knockout_state(raw: object | None = None) -> dict[str, Any]:
    """Return a complete knockout state with schedule plus previous live/admin data."""
    raw_matches = {}
    if isinstance(raw, dict) and isinstance(raw.get("matches"), list):
        for item in raw["matches"]:
            if isinstance(item, dict) and item.get("id") is not None:
                try:
                    raw_matches[int(item["id"])] = item
                except (TypeError, ValueError):
                    continue

    matches: list[dict[str, Any]] = []
    for seed in KNOCKOUT_SCHEDULE:
        prev = raw_matches.get(int(seed["id"]), {})
        match = {
            "id": seed["id"],
            "roundId": seed["roundId"],
            "roundLabel": seed["roundLabel"],
            "kickoffAt": seed["kickoffAt"],
            "homeSlot": seed["homeSlot"],
            "awaySlot": seed["awaySlot"],
            "home": "",
            "away": "",
            "homeScore": None,
            "awayScore": None,
            "isLive": False,
            "isLocked": False,
            "winner": "",
            "nextMatchId": seed.get("nextMatchId"),
            "isScoring": seed.get("isScoring", True),
            "apiSource": "",
            "apiEventId": "",
            "apiHome": "",
            "apiAway": "",
            "apiKickoffAt": "",
            "apiState": "",
            "apiHomeScore": None,
            "apiAwayScore": None,
        }
        for key in ("home", "away", "winner", "apiSource", "apiEventId", "apiHome", "apiAway", "apiKickoffAt", "apiState"):
            if isinstance(prev.get(key), str):
                match[key] = prev[key].strip()
        for key in ("homeScore", "awayScore", "apiHomeScore", "apiAwayScore"):
            value = prev.get(key)
            if value is not None:
                try:
                    match[key] = int(value)
                except (TypeError, ValueError):
                    match[key] = None
        match["isLive"] = bool(prev.get("isLive"))
        match["isLocked"] = bool(prev.get("isLocked"))
        matches.append(match)

    state = deepcopy(raw) if isinstance(raw, dict) else {}
    state["matches"] = matches
    state.setdefault("lastUpdatedAt", None)
    scoring_applied = state.get("scoringApplied")
    if not isinstance(scoring_applied, dict):
        scoring_applied = {}
    scoring_applied.setdefault("r32", False)
    state["scoringApplied"] = scoring_applied
    return state


def _canonical_api_fixture_label(name: str) -> str:
    text = str(name or "").strip()
    if not text:
        return ""
    third = re.match(r"^Third Place Group ([A-L](?:/[A-L])*)$", text, re.I)
    if third:
        return f"Best 3rd {third.group(1).upper()}"
    winner = re.match(r"^Group ([A-L]) Winner$", text, re.I)
    if winner:
        return f"Winner Group {winner.group(1).upper()}"
    runner_up = re.match(r"^Group ([A-L]) 2nd Place$", text, re.I)
    if runner_up:
        return f"Runner-up Group {runner_up.group(1).upper()}"
    return ESPN_TEAM_LABEL_ALIASES.get(text, text)


def is_placeholder_fixture_team(team: str | None) -> bool:
    text = str(team or "").strip().lower()
    if not text:
        return True
    return (
        text.startswith("winner ")
        or text.startswith("loser ")
        or text.startswith("runner-up group ")
        or text.startswith("winner group ")
        or text.startswith("best 3rd ")
        or " winner" in text
        or " loser" in text
        or ("group " in text and ("winner" in text or "place" in text))
    )


def _event_match_key(event: EspnMatch) -> tuple[datetime, str]:
    return parse_iso(event.kickoff_at) or datetime.max.replace(tzinfo=timezone.utc), event.espn_event_id


def _match_for_espn_event(
    knockout: dict[str, Any],
    event: EspnMatch,
    used_match_ids: set[int],
) -> dict[str, Any] | None:
    for match in knockout.get("matches", []):
        if int(match.get("id") or -1) in used_match_ids:
            continue
        if match.get("apiEventId") and str(match.get("apiEventId")) == event.espn_event_id:
            return match

    event_kickoff = parse_iso(event.kickoff_at)
    if event_kickoff is None:
        return None

    candidates: list[tuple[float, int, dict[str, Any]]] = []
    for match in knockout.get("matches", []):
        match_id = int(match.get("id") or -1)
        if match_id in used_match_ids:
            continue
        kickoff = parse_iso(match.get("kickoffAt")) or parse_iso(match.get("apiKickoffAt"))
        if kickoff is None:
            continue
        delta = abs((event_kickoff - kickoff).total_seconds())
        if delta <= 90 * 60:
            candidates.append((delta, match_id, match))
    if not candidates:
        return None
    candidates.sort(key=lambda item: (item[0], item[1]))
    return candidates[0][2]


def sync_knockout_fixtures_from_espn(
    knockout: dict[str, Any],
    *,
    dates: str = ESPN_KNOCKOUT_DATES,
) -> dict[str, Any]:
    payload = fetch_scoreboard(dates=dates)
    events = sorted(parse_espn_events(payload), key=_event_match_key)
    used: set[int] = set()
    updates = 0
    matched = 0
    for event in events:
        match = _match_for_espn_event(knockout, event, used)
        if match is None:
            continue
        matched += 1
        match_id = int(match["id"])
        used.add(match_id)
        home = _canonical_api_fixture_label(event.home)
        away = _canonical_api_fixture_label(event.away)
        kickoff = event.kickoff_at or ""
        previous = (
            match.get("apiEventId"),
            match.get("apiHome"),
            match.get("apiAway"),
            match.get("apiKickoffAt"),
            match.get("apiState"),
            match.get("apiHomeScore"),
            match.get("apiAwayScore"),
            match.get("home"),
            match.get("away"),
            match.get("kickoffAt"),
        )
        match["apiSource"] = "ESPN"
        match["apiEventId"] = event.espn_event_id
        match["apiHome"] = home
        match["apiAway"] = away
        match["apiKickoffAt"] = kickoff
        match["apiState"] = event.state
        match["apiHomeScore"] = event.home_score
        match["apiAwayScore"] = event.away_score
        if kickoff:
            match["kickoffAt"] = kickoff
        should_apply_api_teams = (
            not match.get("winner")
            and (
                not match.get("isLocked")
                or is_placeholder_fixture_team(match.get("home"))
                or is_placeholder_fixture_team(match.get("away"))
                or (
                    normalize_team_name(str(match.get("home") or "")) == normalize_team_name(str(previous[1] or ""))
                    and normalize_team_name(str(match.get("away") or "")) == normalize_team_name(str(previous[2] or ""))
                )
            )
        )
        if should_apply_api_teams:
            match["home"] = home
            match["away"] = away
        match["isLocked"] = bool(
            match.get("home")
            and match.get("away")
            and not is_placeholder_fixture_team(match.get("home"))
            and not is_placeholder_fixture_team(match.get("away"))
        )
        current = (
            match.get("apiEventId"),
            match.get("apiHome"),
            match.get("apiAway"),
            match.get("apiKickoffAt"),
            match.get("apiState"),
            match.get("apiHomeScore"),
            match.get("apiAwayScore"),
            match.get("home"),
            match.get("away"),
            match.get("kickoffAt"),
        )
        if current != previous:
            updates += 1
    return {
        "source": "ESPN",
        "dates": dates,
        "eventsCount": len(events),
        "matchedCount": matched,
        "updatesCount": updates,
    }


def validate_knockout_fixture_lock(match: dict[str, Any], home: str, away: str) -> None:
    if is_placeholder_fixture_team(home) or is_placeholder_fixture_team(away):
        raise ValueError("Fixture still has placeholder teams. Fill actual teams before locking.")


def read_actual_qualifiers(ws: openpyxl.worksheet.worksheet.Worksheet) -> dict[str, list[str]]:
    return {
        round_id: [
            team
            for cell in cells
            if (team := _cell_text(ws[cell].value))
        ]
        for round_id, cells in ROUND_ACTUAL_CELLS.items()
    }


def round_defs() -> list[dict[str, Any]]:
    return [
        {
            "id": item["id"],
            "label": item["label"],
            "expected": item["expected"],
            "points": item["points"],
        }
        for item in KNOCKOUT_ROUNDS
    ]


def read_player_knockout_picks(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    actual: dict[str, list[str]],
) -> list[dict[str, Any]]:
    picks: list[dict[str, Any]] = []
    for round_def in KNOCKOUT_ROUNDS:
        round_id = str(round_def["id"])
        actual_set = {team for team in actual.get(round_id, []) if team}
        teams = []
        points = 0
        for cell in PLAYER_KNOCKOUT_PICK_CELLS[round_id]:
            team = _cell_text(ws[cell].value)
            if not team:
                continue
            is_correct = team in actual_set if actual_set else False
            if is_correct:
                points += int(round_def["points"])
            teams.append(
                {
                    "team": team,
                    "isCorrect": is_correct,
                    "points": int(round_def["points"]) if is_correct else 0,
                }
            )
        picks.append(
            {
                "roundId": round_id,
                "label": round_def["label"],
                "pointsPerTeam": int(round_def["points"]),
                "points": points,
                "teams": teams,
            }
        )
    return picks


def live_qualifier(match: dict[str, Any]) -> str:
    home = match.get("home")
    away = match.get("away")
    home_score = match.get("homeScore")
    away_score = match.get("awayScore")
    if not home or not away or home_score is None or away_score is None:
        return ""
    if int(home_score) > int(away_score):
        return str(home)
    if int(away_score) > int(home_score):
        return str(away)
    return ""


def apply_live_knockout_points(payload: dict[str, Any]) -> None:
    """Add temporary live knockout points and resort leaderboard."""
    knockout = normalize_knockout_state(payload.get("knockout"))
    live_by_round: dict[str, set[str]] = {}
    for match in knockout["matches"]:
        if not match.get("isLive") or not match.get("isScoring", True):
            continue
        qualifier = live_qualifier(match)
        next_round = QUALIFIER_ROUND_BY_MATCH_ROUND.get(str(match.get("roundId")))
        if qualifier and next_round:
            live_by_round.setdefault(next_round, set()).add(qualifier)

    if not live_by_round:
        return

    for entry in payload.get("leaderboard", []):
        live_points = 0
        for round_pick in entry.get("knockoutPicks") or []:
            round_id = str(round_pick.get("roundId") or "")
            live_teams = live_by_round.get(round_id, set())
            if not live_teams:
                continue
            points_per_team = int(round_pick.get("pointsPerTeam") or KNOCKOUT_POINTS.get(round_id, 0))
            for team_pick in round_pick.get("teams") or []:
                if team_pick.get("team") in live_teams and not team_pick.get("isCorrect"):
                    team_pick["isLiveCorrect"] = True
                    live_points += points_per_team
        entry["confirmedPoints"] = entry.get("points", 0)
        entry["liveKnockoutPoints"] = live_points
        entry["points"] = round(float(entry.get("points") or 0) + live_points, 2)

    payload["leaderboard"].sort(key=lambda item: (-float(item.get("points") or 0), str(item.get("name") or "")))
    previous_points: float | None = None
    previous_rank = 0
    for index, entry in enumerate(payload["leaderboard"], start=1):
        points = float(entry.get("points") or 0)
        rank = previous_rank if previous_points is not None and points == previous_points else index
        entry["rank"] = rank
        entry["rankLabel"] = str(rank)
        previous_rank = rank
        previous_points = points


def update_scoring_constants(xlsx_path: Path = XLSX_PATH) -> None:
    """Set workbook scoring constants to the published rules."""
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb["Calc"]
    values = {
        "B177": "Round of 32",
        "C177": 5,
        "E177": 5,
        "B178": "Round of 16",
        "C178": 8,
        "E178": 8,
        "B179": "Quarterfinals",
        "C179": 18,
        "E179": 18,
        "B180": "Semi",
        "C180": 25,
        "E180": 25,
        "B181": "Final",
        "C181": 30,
        "E181": 30,
        "B182": "Winner",
        "C182": 40,
        "E182": 40,
    }
    for cell, value in values.items():
        ws[cell].value = value
    for row in range(149, 157):
        _replace_formula_ref(ws, row, "$E$178", "$E$179")
    for row in range(159, 163):
        _replace_formula_ref(ws, row, "$E$179", "$E$180")
    for row in range(165, 167):
        _replace_formula_ref(ws, row, "$E$180", "$E$181")
    _replace_formula_ref(ws, 169, "$E$181", "$E$182")
    wb.save(xlsx_path)


def _replace_formula_ref(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    row: int,
    old: str,
    new: str,
) -> None:
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row, col)
        value = cell.value
        if isinstance(value, str) and value.startswith("=") and old in value:
            cell.value = value.replace(old, new)


def set_actual_qualifier(
    team: str,
    round_id: str,
    index: int,
    xlsx_path: Path = XLSX_PATH,
) -> None:
    cells = ROUND_ACTUAL_CELLS.get(round_id)
    if not cells:
        raise ValueError(f"Unknown knockout round: {round_id}")
    if index < 0 or index >= len(cells):
        raise ValueError(f"Qualifier index {index} out of range for {round_id}")
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb[SUMMARY]
    ws[cells[index]].value = team
    wb.save(xlsx_path)


def clear_actual_qualifier(
    round_id: str,
    index: int,
    xlsx_path: Path = XLSX_PATH,
) -> None:
    cells = ROUND_ACTUAL_CELLS.get(round_id)
    if not cells:
        raise ValueError(f"Unknown knockout round: {round_id}")
    if index < 0 or index >= len(cells):
        raise ValueError(f"Qualifier index {index} out of range for {round_id}")
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb[SUMMARY]
    ws[cells[index]].value = None
    wb.save(xlsx_path)


def qualifier_target_for_match(match_id: int) -> tuple[str, int] | None:
    match = next((item for item in KNOCKOUT_SCHEDULE if int(item["id"]) == match_id), None)
    if not match or not match.get("isScoring", True):
        return None
    round_id = QUALIFIER_ROUND_BY_MATCH_ROUND.get(str(match["roundId"]))
    if not round_id:
        return None
    round_matches = [
        item for item in KNOCKOUT_SCHEDULE
        if QUALIFIER_ROUND_BY_MATCH_ROUND.get(str(item["roundId"])) == round_id
    ]
    round_matches.sort(key=lambda item: int(item["id"]))
    return round_id, [int(item["id"]) for item in round_matches].index(match_id)


def state_timestamp() -> str:
    return datetime.now(timezone.utc).isoformat()
