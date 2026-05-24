#!/usr/bin/env python3
"""Write actual match score to Summary sheet columns L/M."""

from __future__ import annotations

import argparse
from pathlib import Path

import openpyxl

from scripts.paths import XLSX_PATH

ROOT = Path(__file__).resolve().parent.parent
DEFAULT_XLSX = XLSX_PATH
SUMMARY = "Summary"
MATCH_ROW_START = 4
MATCH_ROW_END = 120


def find_match_row(
    ws: openpyxl.worksheet.worksheet.Worksheet, match_id: int
) -> int:
    """Return Summary row number for the given match id (column J)."""
    for row in range(MATCH_ROW_START, MATCH_ROW_END):
        value = ws[f"J{row}"].value
        if value is not None and int(value) == match_id:
            teams = ws[f"K{row}"].value
            if teams and "-" in str(teams):
                return row
    raise ValueError(f"Match id {match_id} not found on Summary sheet")


def patch_match(
    match_id: int,
    home_score: int,
    away_score: int,
    xlsx_path: Path = DEFAULT_XLSX,
) -> tuple[str, int, int]:
    """Patch Summary L/M for match_id. Returns teams label and scores."""
    xlsx_path = xlsx_path.resolve()
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb[SUMMARY]
    row = find_match_row(ws, match_id)
    teams = str(ws[f"K{row}"].value)
    ws[f"L{row}"].value = home_score
    ws[f"M{row}"].value = away_score
    wb.save(xlsx_path)
    return teams, home_score, away_score


def main() -> None:
    parser = argparse.ArgumentParser(description="Patch match score in xlsx Summary L/M")
    parser.add_argument("match_id", type=int, help="Match number (Summary column J)")
    parser.add_argument("home_score", type=int, help="Home team score (Summary column L)")
    parser.add_argument("away_score", type=int, help="Away team score (Summary column M)")
    parser.add_argument(
        "--xlsx",
        type=Path,
        default=DEFAULT_XLSX,
        help="Path to Master WorldCup26.xlsx",
    )
    args = parser.parse_args()
    teams, home, away = patch_match(
        args.match_id, args.home_score, args.away_score, args.xlsx
    )
    print(f"Match {args.match_id}: {teams} → {home}-{away}")


if __name__ == "__main__":
    main()
