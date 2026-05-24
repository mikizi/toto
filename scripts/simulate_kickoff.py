#!/usr/bin/env python3
"""Simulate tournament kickoff by moving match 1 kickoff into the near future."""

from __future__ import annotations

import argparse
import json
import shutil
import subprocess
import sys
from datetime import datetime, timedelta, timezone
from pathlib import Path

import openpyxl

from scripts.paths import BACKUP_PATH, LATEST_PATH, ROOT, XLSX_PATH
from scripts.reset_scores import reset_scores

SCHEDULE_ROW_START = 7
SCHEDULE_MATCH_ID_COL = 1
SCHEDULE_KICKOFF_COL = 18


def _soffice_available() -> bool:
    return shutil.which("soffice") is not None


def _first_user_sheet(wb: openpyxl.Workbook) -> str | None:
    return next((name for name in wb.sheetnames if name.startswith("001_")), None)


def _find_schedule_row(ws: openpyxl.worksheet.worksheet.Worksheet, match_id: int) -> int:
    for row in range(SCHEDULE_ROW_START, 200):
        value = ws.cell(row, SCHEDULE_MATCH_ID_COL).value
        if value is not None and int(value) == match_id:
            return row
    raise ValueError(f"Match {match_id} not found on schedule sheet")


def kickoff_iso(minutes: float) -> str:
    """UTC ISO string for now + delay."""
    moment = datetime.now(timezone.utc) + timedelta(minutes=minutes)
    return moment.replace(microsecond=0).isoformat()


def patch_json_kickoff(minutes: float, match_id: int = 1) -> str:
    """Patch latest.json kickoff for match_id. Returns new kickoff ISO."""
    if not LATEST_PATH.exists():
        raise FileNotFoundError(LATEST_PATH)

    payload = json.loads(LATEST_PATH.read_text(encoding="utf-8"))
    iso = kickoff_iso(minutes)
    updated = False
    for match in payload.get("matches", []):
        if int(match.get("id", -1)) == match_id:
            match["kickoffAt"] = iso
            updated = True
            break
    if not updated:
        raise ValueError(f"Match {match_id} not found in {LATEST_PATH}")

    LATEST_PATH.write_text(json.dumps(payload, indent=2) + "\n", encoding="utf-8")
    return iso


def patch_xlsx_kickoff(minutes: float, match_id: int = 1, xlsx_path: Path = XLSX_PATH) -> str:
    """Set schedule kickoff in xlsx and return ISO kickoff time."""
    iso = kickoff_iso(minutes)
    kickoff_dt = datetime.fromisoformat(iso.replace("Z", "+00:00"))

    wb = openpyxl.load_workbook(xlsx_path)
    sheet_name = _first_user_sheet(wb)
    if not sheet_name:
        raise RuntimeError("No 001_* user sheet found in xlsx")

    ws = wb[sheet_name]
    row = _find_schedule_row(ws, match_id)
    ws.cell(row, SCHEDULE_KICKOFF_COL).value = kickoff_dt.replace(tzinfo=None)
    wb.save(xlsx_path)
    return iso


def recalc_and_export(xlsx_path: Path = XLSX_PATH) -> None:
    """Recalculate xlsx formulas and export public/data/latest.json."""
    if not _soffice_available():
        raise RuntimeError("LibreOffice (soffice) is required for full simulation")

    subprocess.run(
        [sys.executable, str(ROOT / "scripts/libreoffice_recalc.py"), str(xlsx_path)],
        check=True,
    )
    subprocess.run(
        [sys.executable, str(ROOT / "scripts/export_summary.py"), "--xlsx", str(xlsx_path)],
        check=True,
    )


def setup_simulation(minutes: float, *, json_only: bool, match_id: int = 1) -> str:
    """Prepare day-zero state with match kickoff moved forward."""
    if json_only:
        return patch_json_kickoff(minutes, match_id)

    if not BACKUP_PATH.exists():
        shutil.copy2(XLSX_PATH, BACKUP_PATH)
        print(f"Backup saved → {BACKUP_PATH}")

    reset_scores(xlsx_path=XLSX_PATH, backup=False)
    iso = patch_xlsx_kickoff(minutes, match_id)
    recalc_and_export()
    return iso


def restore_simulation() -> None:
    """Restore xlsx from simulation backup and re-export JSON."""
    if not BACKUP_PATH.exists():
        raise FileNotFoundError(f"No backup at {BACKUP_PATH}")

    shutil.copy2(BACKUP_PATH, XLSX_PATH)
    print(f"Restored xlsx from {BACKUP_PATH}")
    if _soffice_available():
        recalc_and_export()
    else:
        subprocess.run(
            [sys.executable, str(ROOT / "scripts/export_summary.py")],
            check=True,
        )


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Simulate first-match kickoff for local or CI testing"
    )
    parser.add_argument(
        "--minutes",
        type=float,
        default=5.0,
        help="Kickoff delay from now (default: 5)",
    )
    parser.add_argument(
        "--seconds",
        type=float,
        help="Kickoff delay in seconds (overrides --minutes when set)",
    )
    parser.add_argument(
        "--match-id",
        type=int,
        default=1,
        help="Match number to move (default: 1)",
    )
    parser.add_argument(
        "--json-only",
        action="store_true",
        help="Patch latest.json only (no xlsx changes; for CI)",
    )
    parser.add_argument(
        "--restore",
        action="store_true",
        help="Restore xlsx from simulation backup",
    )
    args = parser.parse_args()

    if args.restore:
        restore_simulation()
        return

    delay_minutes = args.minutes
    if args.seconds is not None:
        delay_minutes = args.seconds / 60.0

    iso = setup_simulation(delay_minutes, json_only=args.json_only, match_id=args.match_id)
    delay_label = (
        f"{args.seconds:g} seconds"
        if args.seconds is not None
        else f"{delay_minutes:g} minutes"
    )
    print(f"Match {args.match_id} kickoff set to {iso} ({delay_label} from now)")
    print("")
    print("Next steps:")
    print("  make serve")
    print("  Open http://localhost:8080/ — coming soon until kickoff, then scoreboard")
    if not args.json_only:
        print("  Restore real schedule: python scripts/simulate_kickoff.py --restore")


if __name__ == "__main__":
    main()
