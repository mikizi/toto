#!/usr/bin/env python3
"""Verify patch → recalc → export updates points correctly (simulation)."""

from __future__ import annotations

import shutil
import sys
import tempfile
import time
import unittest
from dataclasses import dataclass
from pathlib import Path

import openpyxl

from scripts.export_summary import build_export, export, write_export
from scripts.libreoffice_recalc import recalc
from scripts.patch_match import patch_match
from scripts.paths import BACKUP_PATH, LATEST_PATH, XLSX_PATH
from scripts.reset_scores import reset_scores
from scripts.validate_export import validate

REAL_USERS = ["MikiZiso3", "MikiZiso2", "Miki_Ziso", "Nir1", "Nir2", "Nir3"]

# Expected Summary points for real users after match 1 (Mexico vs South Africa).
EXPECTED_MATCH_1: dict[tuple[int, int], dict[str, float]] = {
    (1, 0): {
        "MikiZiso3": 5.0,
        "MikiZiso2": 5.0,
        "Miki_Ziso": 5.0,
        "Nir1": 0.0,
        "Nir2": 0.0,
        "Nir3": 3.0,
    },
    (0, 1): {
        "MikiZiso3": 0.0,
        "MikiZiso2": 0.0,
        "Miki_Ziso": 0.0,
        "Nir1": 3.0,
        "Nir2": 5.0,
        "Nir3": 0.0,
    },
}


@dataclass
class StepResult:
    """One simulation step outcome."""

    label: str
    home: int
    away: int
    points: dict[str, float]
    passed: bool
    detail: str


def read_real_user_points(xlsx_path: Path) -> dict[str, float]:
    """Read cached points for real users from Summary F."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["Summary"]
    points: dict[str, float] = {}
    for row in range(79, 149):
        name = ws[f"D{row}"].value
        if name in REAL_USERS:
            val = ws[f"F{row}"].value
            if val is None:
                raise RuntimeError(f"Calc not cached for {name} — recalc failed")
            points[str(name)] = float(val)
    return points


def points_from_export(payload: dict) -> dict[str, float]:
    """Extract real-user points from export JSON."""
    out: dict[str, float] = {}
    for entry in payload.get("leaderboard", []):
        if entry.get("name") in REAL_USERS:
            out[str(entry["name"])] = float(entry.get("points") or 0)
    return out


def _assert_points(actual: dict[str, float], expected: dict[str, float], tol: float = 0.05) -> None:
    for name in REAL_USERS:
        got = actual.get(name, 0.0)
        want = expected.get(name, 0.0)
        if abs(got - want) > tol:
            raise AssertionError(f"{name}: expected {want} pts, got {got}")


def apply_match_result(
    xlsx_path: Path,
    match_id: int,
    home: int,
    away: int,
) -> dict[str, float]:
    """Patch L/M, recalc, return real-user points."""
    patch_match(match_id, home, away, xlsx_path)
    recalc(xlsx_path)
    return read_real_user_points(xlsx_path)


def run_score_simulation(*, apply: bool = False, quiet: bool = False) -> list[StepResult]:
    """
    Run full score-change simulation on a temp copy of the backup xlsx.

    When apply=True, writes the final state to XLSX_PATH and public/data/latest.json.
    """
    if not BACKUP_PATH.exists():
        raise FileNotFoundError(f"Missing simulation backup: {BACKUP_PATH}")

    results: list[StepResult] = []
    with tempfile.TemporaryDirectory() as tmp:
        work = Path(tmp) / "score_sim.xlsx"
        shutil.copy2(BACKUP_PATH, work)
        reset_scores(xlsx_path=work, backup=False)
        recalc(work)

        try:
            day_zero = read_real_user_points(work)
            _assert_points(day_zero, {name: 0.0 for name in REAL_USERS})
            results.append(
                StepResult("Day zero", 0, 0, day_zero, True, "All real users at 0 pts")
            )
        except (AssertionError, RuntimeError) as exc:
            results.append(StepResult("Day zero", 0, 0, {}, False, str(exc)))
            return results

        steps: list[tuple[str, int, int, dict[str, float] | None]] = [
            ("Match 1: Mexico 1-0", 1, 0, EXPECTED_MATCH_1[(1, 0)]),
            ("Match 1: Mexico 0-1", 0, 1, EXPECTED_MATCH_1[(0, 1)]),
            ("Match 1: back to 1-0", 1, 0, EXPECTED_MATCH_1[(1, 0)]),
        ]

        for label, home, away, expected in steps:
            try:
                points = apply_match_result(work, 1, home, away)
                assert expected is not None
                _assert_points(points, expected)
                payload = build_export(work)
                export_pts = points_from_export(payload)
                _assert_points(export_pts, expected)
                errors = validate(payload)
                if errors:
                    raise AssertionError(f"Export validation: {errors}")
                results.append(StepResult(label, home, away, points, True, "Recalc + export OK"))
            except (AssertionError, RuntimeError) as exc:
                results.append(StepResult(label, home, away, {}, False, str(exc)))
                return results

        if apply:
            shutil.copy2(work, XLSX_PATH)
            write_export(build_export(work))
            if not quiet:
                print(f"Applied → {XLSX_PATH}")
                print(f"Applied → {LATEST_PATH}")

    return results


def print_report(results: list[StepResult]) -> bool:
    """Print simulation report. Returns True if all steps passed."""
    print("Score simulation — patch → recalc → export")
    print("=" * 52)
    all_ok = True
    for step in results:
        status = "PASS" if step.passed else "FAIL"
        print(f"[{status}] {step.label} ({step.home}-{step.away})")
        if step.passed and step.points:
            highlights = [
                f"{name}={step.points[name]:.0f}"
                for name in ("MikiZiso3", "Nir2", "Nir1")
                if name in step.points
            ]
            print(f"       {', '.join(highlights)}")
        elif not step.passed:
            print(f"       {step.detail}")
            all_ok = False
    print("=" * 52)
    if all_ok:
        print("All score-change checks passed.")
    else:
        print("Score simulation FAILED.")
    return all_ok


def main() -> int:
    import argparse

    parser = argparse.ArgumentParser(
        description="Simulate match score changes and verify points update"
    )
    parser.add_argument(
        "--apply",
        action="store_true",
        help="Write final xlsx + latest.json after successful simulation",
    )
    parser.add_argument(
        "--apply-result",
        metavar="H-A",
        help="After simulation, apply one result to main xlsx (e.g. 0-1)",
    )
    args = parser.parse_args()

    try:
        results = run_score_simulation(apply=args.apply and not args.apply_result)
    except FileNotFoundError as exc:
        print(exc, file=sys.stderr)
        return 1

    ok = print_report(results)
    if not ok:
        return 1

    if args.apply_result:
        try:
            home_s, away_s = args.apply_result.split("-", 1)
            home, away = int(home_s), int(away_s)
        except ValueError:
            print("Use --apply-result 0-1", file=sys.stderr)
            return 1
        if not BACKUP_PATH.exists():
            shutil.copy2(XLSX_PATH, BACKUP_PATH)
        reset_scores(xlsx_path=XLSX_PATH, backup=False)
        points = apply_match_result(XLSX_PATH, 1, home, away)
        export(XLSX_PATH)
        print(f"\nApplied {home}-{away} to main xlsx + latest.json")
        for name in ("Nir2", "Nir1", "MikiZiso3"):
            print(f"  {name}: {points.get(name, 0):.0f} pts")

    return 0


class TestScoreSimulation(unittest.TestCase):
    """Integration: score patches must update Calc and export."""

    @classmethod
    def setUpClass(cls) -> None:
        if not BACKUP_PATH.exists():
            raise unittest.SkipTest("simulation backup not found")
        if not shutil.which("soffice"):
            raise unittest.SkipTest("LibreOffice (soffice) not installed")

    def test_score_change_simulation(self) -> None:
        results = run_score_simulation()
        failed = [step for step in results if not step.passed]
        self.assertEqual(failed, [], msg=str(failed))


if __name__ == "__main__":
    sys.exit(main())
