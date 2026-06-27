#!/usr/bin/env python3
"""Export Summary sheet from recalculated xlsx to data/latest.json."""

from __future__ import annotations

import argparse
import json
import shutil
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import openpyxl

from scripts.live_state import DEFAULT_BROADCAST, normalize_broadcast
from scripts.knockout import (
    KNOCKOUT_ROUNDS,
    QUALIFIER_ROUND_BY_MATCH_ROUND,
    apply_live_knockout_points,
    is_placeholder_fixture_team,
    normalize_knockout_state,
    read_actual_qualifiers,
    read_player_knockout_picks,
    round_defs,
)
from scripts.paths import DATA_DIR, LATEST_PATH, VERSION_PATH, VERSIONS_DIR, XLSX_PATH
from scripts.registration import normalize_registration

DEFAULT_XLSX = XLSX_PATH

SUMMARY = "Summary"
USER_ROW_START = 79
USER_ROW_END = 200
SUMMARY_LEADERBOARD_ROW_START = 4
SUMMARY_LEADERBOARD_ROW_END = 80
CALC_SCORE_ROW_START = 238
MATCH_ROW_START = 4
MATCH_COUNT = 72
MATCH_ROW_END = MATCH_ROW_START + MATCH_COUNT
CHAMPION_CELL = "CM47"
SCHEDULE_ROW_START = 7
SCHEDULE_MATCH_ID_COL = 1
SCHEDULE_KICKOFF_COL = 18
PICK_HOME_COL = 6
PICK_AWAY_COL = 7
MATCH_RESULT_POINTS = 3.0
EXACT_SCORE_POINTS = 2.0
LATE_JOINER_PREFIX = "N_"
LATE_JOINER_BASELINE_MATCH_COUNT = 24

def _kickoff_to_iso(value: object) -> str | None:
    """Convert Excel kickoff cell to UTC ISO-8601 string."""
    if value is None:
        return None
    if isinstance(value, datetime):
        dt = value
    elif isinstance(value, (int, float)):
        from datetime import timedelta

        base = datetime(1899, 12, 30)
        days = int(value)
        frac = value - days
        dt = base + timedelta(days=days, seconds=round(frac * 86400))
    else:
        return None
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt.isoformat()


def _read_match_kickoffs(wb: openpyxl.Workbook) -> dict[int, str]:
    """Read kickoff times from the first user sheet schedule (col A + col R)."""
    kickoffs: dict[int, str] = {}
    schedule_sheet = next((name for name in wb.sheetnames if name.startswith("001_")), None)
    if not schedule_sheet:
        return kickoffs

    ws = wb[schedule_sheet]
    for row in range(SCHEDULE_ROW_START, 200):
        match_id = ws.cell(row, SCHEDULE_MATCH_ID_COL).value
        kickoff = ws.cell(row, SCHEDULE_KICKOFF_COL).value
        if match_id is None or kickoff is None:
            continue
        try:
            mid = int(match_id)
        except (TypeError, ValueError):
            continue
        iso = _kickoff_to_iso(kickoff)
        if iso:
            kickoffs[mid] = iso
    return kickoffs


def _user_sheet_name(uid: object, name: str) -> str:
    """Build user sheet name from Summary id + display name."""
    uid_text = str(uid).strip().zfill(3)
    return f"{uid_text}_{name}"


def _user_sheet_for_uid(wb: openpyxl.Workbook, uid: object) -> str | None:
    """Find a user sheet by Summary id, independent of cached display name."""
    uid_text = str(uid or "").strip().zfill(3)
    prefix = f"{uid_text}_"
    return next((name for name in wb.sheetnames if name.startswith(prefix)), None)


def _cell_text(value: object) -> str | None:
    """Parse a cached spreadsheet cell as text, ignoring Excel error strings."""
    if value is None:
        return None
    text = str(value).strip()
    if not text or text.startswith("#"):
        return None
    return text


def _read_champion(
    wb: openpyxl.Workbook,
    ws: openpyxl.worksheet.worksheet.Worksheet,
    row: int,
) -> str | None:
    """Read champion pick from Summary E, falling back to user sheet CM47."""
    champion = _cell_text(ws[f"E{row}"].value)
    if champion:
        return champion

    uid = ws[f"C{row}"].value
    name = ws[f"D{row}"].value
    if not uid or not name:
        return None

    sheet_name = _user_sheet_name(uid, str(name))
    if sheet_name not in wb.sheetnames:
        return None

    return _cell_text(wb[sheet_name][CHAMPION_CELL].value)


def _split_teams(teams: str) -> tuple[str, str]:
    home, away = teams.split("-", 1)
    return home.strip(), away.strip()


def _is_match_row(ws: openpyxl.worksheet.worksheet.Worksheet, row: int) -> bool:
    match_id = ws[f"J{row}"].value
    teams = ws[f"K{row}"].value
    try:
        int(match_id)
    except (TypeError, ValueError):
        return False
    return teams is not None and "-" in str(teams)


def _read_matches(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    kickoffs: dict[int, str] | None = None,
) -> list[dict[str, Any]]:
    kickoffs = kickoffs or {}
    matches: list[dict[str, Any]] = []
    for row in range(MATCH_ROW_START, MATCH_ROW_END):
        if not _is_match_row(ws, row):
            continue
        teams = str(ws[f"K{row}"].value)
        home, away = _split_teams(teams)
        home_score = ws[f"L{row}"].value
        away_score = ws[f"M{row}"].value
        played = home_score is not None and away_score is not None
        match_id = int(ws[f"J{row}"].value)
        matches.append(
            {
                "id": match_id,
                "teams": teams,
                "home": home,
                "away": away,
                "homeScore": int(home_score) if played else None,
                "awayScore": int(away_score) if played else None,
                "played": played,
                "kickoffAt": kickoffs.get(match_id),
            }
        )
    return matches


def _read_cell_range_text(
    ws: openpyxl.worksheet.worksheet.Worksheet, cell_range: str
) -> list[str]:
    """Read non-empty, non-error text values from an Excel range."""
    teams: list[str] = []
    for row in ws[cell_range]:
        for cell in row:
            team = _cell_text(cell.value)
            if team:
                teams.append(team)
    return teams


def _read_knockout(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    previous: dict[str, Any] | None = None,
) -> dict[str, Any]:
    """Read actual knockout qualifiers entered in the Summary sheet."""
    previous_knockout = previous.get("knockout") if previous else None
    knockout = normalize_knockout_state(previous_knockout)
    actual = read_actual_qualifiers(ws)
    scoring_applied = knockout.get("scoringApplied") if isinstance(knockout.get("scoringApplied"), dict) else {}
    scoring_applied["r32"] = bool(scoring_applied.get("r32") or len(actual.get("r32", [])) >= 32)
    knockout.update({
        "rounds": round_defs(),
        "actual": actual,
        "scoringApplied": scoring_applied,
        "points": {
            str(round_def["id"]): round_def["points"]
            for round_def in KNOCKOUT_ROUNDS
        },
    })
    return knockout


def _tournament_teams(matches: list[dict[str, Any]]) -> list[str]:
    teams: dict[str, str] = {}
    for match in matches:
        for side in ("home", "away"):
            team = _cell_text(match.get(side))
            if team:
                teams.setdefault(team.casefold(), team)
    return sorted(teams.values(), key=str.casefold)


def _concrete_team(value: object) -> str | None:
    team = _cell_text(value)
    if not team or is_placeholder_fixture_team(team):
        return None
    return team


def _match_api_winner(match: dict[str, Any]) -> str | None:
    if str(match.get("apiState") or "").lower() != "post":
        return None
    home = _concrete_team(match.get("apiHome") or match.get("home"))
    away = _concrete_team(match.get("apiAway") or match.get("away"))
    if not home or not away:
        return None
    try:
        home_score = int(match.get("apiHomeScore"))
        away_score = int(match.get("apiAwayScore"))
    except (TypeError, ValueError):
        return None
    if home_score == away_score:
        return None
    return home if home_score > away_score else away


def _match_winner(match: dict[str, Any]) -> str | None:
    return _concrete_team(match.get("winner")) or _match_api_winner(match)


def _match_loser(match: dict[str, Any], winner: str) -> str | None:
    home = _concrete_team(match.get("home") or match.get("apiHome"))
    away = _concrete_team(match.get("away") or match.get("apiAway"))
    if home == winner and away:
        return away
    if away == winner and home:
        return home
    return None


def _round_qualifier_sets(knockout: dict[str, Any]) -> dict[str, set[str]]:
    actual = knockout.get("actual") if isinstance(knockout.get("actual"), dict) else {}
    qualifiers: dict[str, set[str]] = {
        str(round_def["id"]): {
            team
            for value in actual.get(str(round_def["id"]), [])
            if (team := _concrete_team(value))
        }
        for round_def in KNOCKOUT_ROUNDS
    }
    for match in knockout.get("matches") or []:
        if not isinstance(match, dict):
            continue
        if match.get("roundId") == "r32_match":
            for side in ("home", "away", "apiHome", "apiAway"):
                team = _concrete_team(match.get(side))
                if team:
                    qualifiers.setdefault("r32", set()).add(team)
        target_round = QUALIFIER_ROUND_BY_MATCH_ROUND.get(str(match.get("roundId") or ""))
        winner = _match_winner(match)
        if target_round and winner:
            qualifiers.setdefault(target_round, set()).add(winner)
    return qualifiers


def _knockout_eliminated(
    knockout: dict[str, Any],
    matches: list[dict[str, Any]],
) -> dict[str, list[str]]:
    qualifiers = _round_qualifier_sets(knockout)
    eliminated: dict[str, set[str]] = {str(round_def["id"]): set() for round_def in KNOCKOUT_ROUNDS}
    tournament_teams = set(_tournament_teams(matches))

    r32 = qualifiers.get("r32", set())
    r32_expected = int(KNOCKOUT_ROUNDS[0]["expected"])
    if len(r32) >= r32_expected:
        eliminated["r32"].update(team for team in tournament_teams if team not in r32)

    for match in knockout.get("matches") or []:
        if not isinstance(match, dict):
            continue
        target_round = QUALIFIER_ROUND_BY_MATCH_ROUND.get(str(match.get("roundId") or ""))
        winner = _match_winner(match)
        if not target_round or not winner:
            continue
        loser = _match_loser(match, winner)
        if loser:
            eliminated.setdefault(target_round, set()).add(loser)

    round_order = [str(round_def["id"]) for round_def in KNOCKOUT_ROUNDS]
    for index in range(1, len(round_order)):
        previous_round = round_order[index - 1]
        round_id = round_order[index]
        current = qualifiers.get(round_id, set())
        expected = int(KNOCKOUT_ROUNDS[index]["expected"])
        if len(current) >= expected:
            eliminated[round_id].update(
                team for team in qualifiers.get(previous_round, set()) if team not in current
            )

    return {
        round_id: sorted(teams, key=str.casefold)
        for round_id, teams in eliminated.items()
        if teams
    }


def _is_test_user(name: str) -> bool:
    """Test sheets use names like test1, test54 — hide from public scoreboard."""
    return name.lower().startswith("test")


def _is_late_joiner(name: str) -> bool:
    """Late entrants get average points for the first round, not historical picks."""
    return name.strip().startswith(LATE_JOINER_PREFIX)


def _public_leaderboard(
    raw: list[dict[str, Any]], previous: dict[str, Any] | None
) -> list[dict[str, Any]]:
    """Pool-only leaderboard in the exact order read from Summary."""
    pool = [entry for entry in raw if not _is_test_user(entry["name"])]
    return _movement(pool, previous)


def _cell_number(value: object) -> float | None:
    """Parse a cached spreadsheet cell value, ignoring Excel error strings."""
    if value is None:
        return None
    if isinstance(value, str):
        text = value.strip()
        if not text or text.startswith("#"):
            return None
        try:
            return float(text)
        except ValueError:
            return None
    if isinstance(value, (int, float)):
        return float(value)
    return None


def _cell_int(value: object) -> int | None:
    """Parse a cached spreadsheet cell as an integer rank."""
    number = _cell_number(value)
    if number is None:
        return None
    return int(round(number))


def _result_key(home: float, away: float) -> str:
    if home > away:
        return "I"
    if home < away:
        return "II"
    return "X"


def _score_prediction(
    actual_home: int,
    actual_away: int,
    pick_home: object,
    pick_away: object,
) -> float:
    home = _cell_number(pick_home)
    away = _cell_number(pick_away)
    if home is None or away is None:
        return 0.0

    points = 0.0
    if _result_key(home, away) == _result_key(actual_home, actual_away):
        points += MATCH_RESULT_POINTS
    if int(home) == actual_home and int(away) == actual_away:
        points += EXACT_SCORE_POINTS
    return points


def _score_from_user_sheet(
    wb: openpyxl.Workbook,
    ws_summary: openpyxl.worksheet.worksheet.Worksheet,
    row: int,
    matches: list[dict[str, Any]],
    baseline_points: float | None = None,
) -> float | None:
    """Compute group-stage points from user picks when cached formulas are missing."""
    uid = ws_summary[f"C{row}"].value
    name = ws_summary[f"D{row}"].value
    if not uid or not name:
        return None

    sheet_name = _user_sheet_name(uid, str(name))
    if sheet_name not in wb.sheetnames:
        return None

    ws_user = wb[sheet_name]
    pick_rows: dict[int, int] = {}
    for user_row in range(SCHEDULE_ROW_START, 200):
        match_id = ws_user.cell(user_row, SCHEDULE_MATCH_ID_COL).value
        try:
            pick_rows[int(match_id)] = user_row
        except (TypeError, ValueError):
            continue

    total = baseline_points if _is_late_joiner(str(name)) and baseline_points is not None else 0.0
    for match in matches:
        if not match["played"]:
            continue
        if _is_late_joiner(str(name)) and int(match["id"]) <= LATE_JOINER_BASELINE_MATCH_COUNT:
            continue
        pick_row = pick_rows.get(int(match["id"]))
        if pick_row is None:
            continue
        total += _score_prediction(
            int(match["homeScore"]),
            int(match["awayScore"]),
            ws_user.cell(pick_row, PICK_HOME_COL).value,
            ws_user.cell(pick_row, PICK_AWAY_COL).value,
        )
    return round(total, 2)


def _score_from_user_sheet_name(
    wb: openpyxl.Workbook,
    sheet_name: str,
    matches: list[dict[str, Any]],
    baseline_points: float | None = None,
) -> float:
    """Compute group-stage points directly from a known user sheet name."""
    ws_user = wb[sheet_name]
    name = sheet_name.split("_", 1)[1] if "_" in sheet_name else sheet_name
    pick_rows: dict[int, int] = {}
    for user_row in range(SCHEDULE_ROW_START, 200):
        match_id = ws_user.cell(user_row, SCHEDULE_MATCH_ID_COL).value
        try:
            pick_rows[int(match_id)] = user_row
        except (TypeError, ValueError):
            continue

    total = baseline_points if _is_late_joiner(name) and baseline_points is not None else 0.0
    for match in matches:
        if not match["played"]:
            continue
        if _is_late_joiner(name) and int(match["id"]) <= LATE_JOINER_BASELINE_MATCH_COUNT:
            continue
        pick_row = pick_rows.get(int(match["id"]))
        if pick_row is None:
            continue
        total += _score_prediction(
            int(match["homeScore"]),
            int(match["awayScore"]),
            ws_user.cell(pick_row, PICK_HOME_COL).value,
            ws_user.cell(pick_row, PICK_AWAY_COL).value,
        )
    return round(total, 2)


def _player_picks_from_user_sheet(
    wb: openpyxl.Workbook,
    sheet_name: str,
    matches: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    """Read one player's predictions for every known match."""
    ws_user = wb[sheet_name]
    name = sheet_name.split("_", 1)[1] if "_" in sheet_name else sheet_name
    pick_rows: dict[int, int] = {}
    for user_row in range(SCHEDULE_ROW_START, 200):
        match_id = ws_user.cell(user_row, SCHEDULE_MATCH_ID_COL).value
        try:
            pick_rows[int(match_id)] = user_row
        except (TypeError, ValueError):
            continue

    picks: list[dict[str, Any]] = []
    for match in matches:
        match_id = int(match["id"])
        pick_row = pick_rows.get(match_id)
        home_pick = away_pick = None
        if pick_row is not None:
            home_pick = _cell_int(ws_user.cell(pick_row, PICK_HOME_COL).value)
            away_pick = _cell_int(ws_user.cell(pick_row, PICK_AWAY_COL).value)

        points = None
        if (
            match["played"]
            and not (
                _is_late_joiner(name)
                and match_id <= LATE_JOINER_BASELINE_MATCH_COUNT
            )
            and home_pick is not None
            and away_pick is not None
            and match["homeScore"] is not None
            and match["awayScore"] is not None
        ):
            points = _score_prediction(
                int(match["homeScore"]),
                int(match["awayScore"]),
                home_pick,
                away_pick,
            )

        picks.append(
            {
                "matchId": match_id,
                "homePick": home_pick,
                "awayPick": away_pick,
                "points": points,
            }
        )
    return picks


def _read_user_points(
    wb_data: openpyxl.Workbook,
    wb_formulas: openpyxl.Workbook,
    ws_data: openpyxl.worksheet.worksheet.Worksheet,
    row: int,
    matches: list[dict[str, Any]] | None = None,
    baseline_points: float | None = None,
) -> float:
    """Read cached Summary points, falling back to Calc when Summary F is empty."""
    cached = _cell_number(ws_data[f"F{row}"].value)
    if cached is not None:
        return round(cached, 2)

    formula = wb_formulas[SUMMARY][f"F{row}"].value
    if isinstance(formula, str) and formula.upper().startswith("=CALC!"):
        calc_ref = formula.split("!", 1)[1].strip()
        calc_val = _cell_number(wb_data["Calc"][calc_ref].value)
        if calc_val is not None:
            return round(calc_val, 2)

    if matches is not None:
        computed = _score_from_user_sheet(
            wb_formulas,
            ws_data,
            row,
            matches,
            baseline_points,
        )
        if computed is not None:
            return computed
    return 0.0


def count_played_matches(xlsx_path: Path) -> int:
    """Count Summary rows with both home and away scores filled in."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[SUMMARY]
    return sum(
        1
        for row in range(MATCH_ROW_START, MATCH_ROW_END)
        if _is_match_row(ws, row)
        and ws[f"L{row}"].value is not None
        and ws[f"M{row}"].value is not None
    )


def assert_recalc_cached(xlsx_path: Path, *, games_played: int) -> None:
    """Fail fast when LibreOffice did not cache formula results."""
    if games_played <= 0:
        return

    wb_data = openpyxl.load_workbook(xlsx_path, data_only=True)
    wb_formulas = openpyxl.load_workbook(xlsx_path, data_only=False)
    ws = wb_data[SUMMARY]
    missing: list[str] = []
    for row in range(USER_ROW_START, USER_ROW_END):
        name = ws[f"D{row}"].value
        if not name or name == "Name" or _is_test_user(str(name)):
            continue
        raw = ws[f"F{row}"].value
        if isinstance(raw, str) and raw.strip().startswith("#"):
            missing.append(str(name))
            continue
        if _cell_number(raw) is None and _read_user_points(
            wb_data, wb_formulas, ws, row
        ) == 0.0:
            missing.append(str(name))

    if missing:
        names = ", ".join(missing[:5])
        extra = f" (+{len(missing) - 5} more)" if len(missing) > 5 else ""
        raise RuntimeError(
            f"Leaderboard points not cached after recalc ({names}{extra}). "
            "LibreOffice did not refresh the workbook."
        )


def _read_leaderboard(
    wb_data: openpyxl.Workbook,
    ws_data: openpyxl.worksheet.worksheet.Worksheet,
    wb_formulas: openpyxl.Workbook,
    matches: list[dict[str, Any]],
    knockout_actual: dict[str, list[str]],
) -> list[dict[str, Any]]:
    raw_by_name = _read_raw_leaderboard_by_name(
        wb_data,
        ws_data,
        wb_formulas,
        matches,
        knockout_actual,
    )
    visible_rows = _read_visible_summary_leaderboard(ws_data, raw_by_name)
    public_raw_count = sum(
        1 for entry in raw_by_name.values() if not _is_test_user(entry["name"])
    )
    if visible_rows and len(visible_rows) >= public_raw_count:
        return visible_rows
    return _reconstructed_summary_leaderboard(list(raw_by_name.values()))


def _summary_raw_header_row(ws: openpyxl.worksheet.worksheet.Worksheet) -> int | None:
    """Find the raw user table header on Summary."""
    max_row = min(ws.max_row + 1, USER_ROW_END)
    for row in range(SUMMARY_LEADERBOARD_ROW_START, max_row):
        marker = ws[f"C{row}"].value
        if str(marker or "").strip() == "#" and _cell_text(ws[f"D{row}"].value) == "Name":
            return row
    return None


def _read_raw_leaderboard_by_name(
    wb_data: openpyxl.Workbook,
    ws_data: openpyxl.worksheet.worksheet.Worksheet,
    wb_formulas: openpyxl.Workbook,
    matches: list[dict[str, Any]],
    knockout_actual: dict[str, list[str]],
) -> dict[str, dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    raw_start = (_summary_raw_header_row(ws_data) or (USER_ROW_START - 1)) + 1
    late_joiner_baseline = _late_joiner_baseline_points(wb_formulas, ws_data, matches)
    for row in range(raw_start, USER_ROW_END):
        uid = ws_data[f"C{row}"].value
        name = _cell_text(ws_data[f"D{row}"].value)
        if not name or name == "Name":
            sheet_name = _user_sheet_for_uid(wb_formulas, uid)
            if sheet_name is None:
                continue
            name = sheet_name.split("_", 1)[1]
        if not name or name == "Name":
            continue
        points = _read_user_points(
            wb_data,
            wb_formulas,
            ws_data,
            row,
            matches,
            late_joiner_baseline,
        )
        if points == 0.0:
            sheet_name = _user_sheet_for_uid(wb_formulas, uid)
            if sheet_name is not None:
                points = _score_from_user_sheet_name(
                    wb_formulas,
                    sheet_name,
                    matches,
                    late_joiner_baseline,
                )
        rank = _cell_int(ws_data[f"G{row}"].value)
        champion = _read_champion(wb_data, ws_data, row)
        if not champion:
            sheet_name = _user_sheet_for_uid(wb_formulas, uid)
            if sheet_name is not None:
                champion = _cell_text(wb_formulas[sheet_name][CHAMPION_CELL].value)
        sheet_name = _user_sheet_for_uid(wb_formulas, uid)
        picks = (
            _player_picks_from_user_sheet(wb_formulas, sheet_name, matches)
            if sheet_name is not None
            else []
        )
        knockout_picks = (
            read_player_knockout_picks(wb_formulas[sheet_name], knockout_actual)
            if sheet_name is not None
            else []
        )
        rows.append(
            {
                "id": str(uid or ""),
                "name": name,
                "points": points,
                "rank": rank,
                "rankLabel": str(rank) if rank is not None else None,
                "champion": champion,
                "picks": picks,
                "knockoutPicks": knockout_picks,
                "knockoutPoints": sum(int(item.get("points") or 0) for item in knockout_picks),
                "summaryOrder": row,
            }
        )
    return {entry["name"]: entry for entry in rows}


def _late_joiner_baseline_points(
    wb: openpyxl.Workbook,
    ws_summary: openpyxl.worksheet.worksheet.Worksheet,
    matches: list[dict[str, Any]],
) -> float | None:
    """Average first-round score for original players, used as late-entry baseline."""
    first_round = [
        match
        for match in matches
        if int(match["id"]) <= LATE_JOINER_BASELINE_MATCH_COUNT and match["played"]
    ]
    if not first_round:
        return None

    scores: list[float] = []
    raw_start = (_summary_raw_header_row(ws_summary) or (USER_ROW_START - 1)) + 1
    for row in range(raw_start, USER_ROW_END):
        uid = ws_summary[f"C{row}"].value
        sheet_name = _user_sheet_for_uid(wb, uid)
        if sheet_name is None:
            continue
        name = sheet_name.split("_", 1)[1]
        if _is_test_user(name) or _is_late_joiner(name):
            continue
        scores.append(_score_from_user_sheet_name(wb, sheet_name, first_round))

    if not scores:
        return None
    return round(sum(scores) / len(scores), 2)


def _read_visible_summary_leaderboard(
    ws_data: openpyxl.worksheet.worksheet.Worksheet,
    raw_by_name: dict[str, dict[str, Any]],
) -> list[dict[str, Any]]:
    """Read the sorted leaderboard table displayed at the top of Summary."""
    rows: list[dict[str, Any]] = []
    raw_header = _summary_raw_header_row(ws_data)
    end_row = (
        raw_header - 1
        if raw_header is not None
        else min(SUMMARY_LEADERBOARD_ROW_END, ws_data.max_row)
    )
    for row in range(SUMMARY_LEADERBOARD_ROW_START, end_row + 1):
        name = _cell_text(ws_data[f"D{row}"].value)
        if not name or name == "Name":
            continue
        name_text = name
        if _is_test_user(name_text):
            continue

        rank_label = _cell_text(ws_data[f"C{row}"].value)
        raw = raw_by_name.get(name_text, {})
        rank = raw.get("rank") or _cell_int(ws_data[f"A{row}"].value)
        points = _cell_number(ws_data[f"F{row}"].value)
        champion = _cell_text(ws_data[f"E{row}"].value)
        if rank_label == "-" and rows:
            rank = rows[-1]["rank"]
        if rank_label == "-" and rank is not None:
            rank_label = str(rank)
        rows.append(
            {
                "id": str(raw.get("id") or ""),
                "name": name_text,
                "points": round(points or 0.0, 2),
                "rank": rank,
                "rankLabel": rank_label,
                "champion": champion or raw.get("champion"),
                "picks": raw.get("picks") or [],
                "knockoutPicks": raw.get("knockoutPicks") or [],
                "knockoutPoints": raw.get("knockoutPoints") or 0,
            }
        )
    return rows


def _reconstructed_summary_leaderboard(raw_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Rebuild the visible Summary order when cached VLOOKUP rows are #N/A."""
    pool = [entry for entry in raw_rows if not _is_test_user(entry["name"])]
    pool.sort(key=lambda e: (-float(e.get("points") or 0), int(e.get("summaryOrder") or 9999)))

    rows: list[dict[str, Any]] = []
    previous_points: float | None = None
    previous_rank: int | None = None
    for index, entry in enumerate(pool, start=1):
        points = float(entry.get("points") or 0)
        rank = entry.get("rank")
        if rank is None:
            rank = previous_rank if previous_points is not None and points == previous_points else index
        rank_label = str(rank)
        previous_points = points
        previous_rank = rank
        public_entry = {k: v for k, v in entry.items() if k != "summaryOrder"}
        public_entry["rank"] = rank
        public_entry["rankLabel"] = rank_label
        rows.append(public_entry)
    return rows


def _effective_rank(entry: dict[str, Any], leaderboard: list[dict[str, Any]], index: int) -> int | None:
    label = str(entry.get("rankLabel") or "").strip()
    if label and label != "-":
        parsed = _cell_int(label)
        if parsed is not None:
            return parsed
    if label == "-":
        for previous in reversed(leaderboard[:index]):
            previous_label = str(previous.get("rankLabel") or "").strip()
            if previous_label and previous_label != "-":
                parsed = _cell_int(previous_label)
                if parsed is not None:
                    return parsed
    rank = entry.get("rank")
    return rank if isinstance(rank, int) else _cell_int(rank)


def _movement(
    current: list[dict[str, Any]], previous: dict[str, Any] | None
) -> list[dict[str, Any]]:
    prev_ranks: dict[str, int | None] = {}
    if previous:
        previous_leaderboard = previous.get("leaderboard", [])
        if isinstance(previous_leaderboard, list):
            for index, entry in enumerate(previous_leaderboard):
                if isinstance(entry, dict):
                    prev_ranks[str(entry.get("name"))] = _effective_rank(
                        entry,
                        previous_leaderboard,
                        index,
                    )

    result: list[dict[str, Any]] = []
    for index, entry in enumerate(current):
        name = entry["name"]
        rank = _effective_rank(entry, current, index)
        prev = prev_ranks.get(name)
        if prev is None or rank is None:
            move = "same"
        elif rank < prev:
            move = "up"
        elif rank > prev:
            move = "down"
        else:
            move = "same"
        result.append({**entry, "movement": move})
    return result


def _last_result(matches: list[dict[str, Any]]) -> dict[str, Any] | None:
    played = [m for m in matches if m["played"]]
    if not played:
        return None
    last = max(played, key=lambda m: m["id"])
    return {
        "matchId": last["id"],
        "teams": last["teams"],
        "home": last["home"],
        "away": last["away"],
        "homeScore": last["homeScore"],
        "awayScore": last["awayScore"],
    }


def build_export(xlsx_path: Path, previous: dict[str, Any] | None = None) -> dict[str, Any]:
    """Read recalculated xlsx and build export payload."""
    wb_data = openpyxl.load_workbook(xlsx_path, data_only=True)
    wb_formulas = openpyxl.load_workbook(xlsx_path, data_only=False)
    ws = wb_data[SUMMARY]
    kickoffs = _read_match_kickoffs(wb_data)
    matches = _read_matches(ws, kickoffs)
    knockout = _read_knockout(ws, previous)
    knockout["eliminated"] = _knockout_eliminated(knockout, matches)
    knockout_actual = knockout.get("actual") if isinstance(knockout.get("actual"), dict) else {}
    raw_leaderboard = _read_leaderboard(
        wb_data,
        ws,
        wb_formulas,
        matches,
        knockout_actual,
    )
    leaderboard = _public_leaderboard(raw_leaderboard, previous)
    version = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H%M%SZ")
    played_count = sum(1 for m in matches if m["played"])
    broadcast = normalize_broadcast(
        previous.get("broadcast") if previous else DEFAULT_BROADCAST
    )
    previous_registration = (previous or {}).get("registration")
    if not isinstance(previous_registration, dict):
        previous_registration = {
            "users": [entry["name"] for entry in leaderboard if entry.get("name")]
        }
    registration = normalize_registration(previous_registration, matches)
    payload = {
        "version": version,
        "generatedAt": datetime.now(timezone.utc).isoformat(),
        "gamesPlayed": played_count,
        "lastResult": _last_result(matches),
        "leaderboard": leaderboard,
        "matches": matches,
        "knockout": knockout,
        "broadcast": broadcast,
        "registration": registration,
    }
    apply_live_knockout_points(payload)
    return payload


def write_export(
    payload: dict[str, Any],
    latest_path: Path = LATEST_PATH,
    version_path: Path = VERSION_PATH,
    versions_dir: Path = VERSIONS_DIR,
) -> Path:
    """Write latest.json, version.json, and a version snapshot."""
    latest_path.parent.mkdir(parents=True, exist_ok=True)
    versions_dir.mkdir(parents=True, exist_ok=True)

    latest_path.write_text(json.dumps(payload, indent=2) + "\n", encoding="utf-8")
    version_path.write_text(
        json.dumps({"version": payload["version"]}, indent=2) + "\n",
        encoding="utf-8",
    )
    snapshot = versions_dir / f"{payload['version']}.json"
    shutil.copy2(latest_path, snapshot)
    return snapshot


def export(xlsx_path: Path = DEFAULT_XLSX) -> dict[str, Any]:
    """Export xlsx to data/latest.json with movement vs previous snapshot."""
    previous: dict[str, Any] | None = None
    if LATEST_PATH.exists():
        previous = json.loads(LATEST_PATH.read_text(encoding="utf-8"))
    payload = build_export(xlsx_path, previous)
    snapshot = write_export(payload)
    print(f"Exported {len(payload['leaderboard'])} users, {payload['gamesPlayed']} games played")
    print(f"→ {LATEST_PATH}")
    print(f"→ {snapshot}")
    return payload


def main() -> None:
    parser = argparse.ArgumentParser(description="Export xlsx Summary to data/latest.json")
    parser.add_argument(
        "--xlsx",
        type=Path,
        default=DEFAULT_XLSX,
        help="Path to Master WorldCup26.xlsx",
    )
    args = parser.parse_args()
    export(args.xlsx)


if __name__ == "__main__":
    main()
